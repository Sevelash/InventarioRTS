from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, jsonify
from models import db, Asset, Category, Client, Employee, Assignment, Shipment, AuditLog, log_action, ALL_MODULES, Department, AccessRequest, Supplier, Brand, IDConfig, PurchaseOrder, Invoice, Maintenance, License, LicenseAssignment, AppSetting, Evaluation, EvaluationGoal, EvaluationCompetency
from datetime import datetime, date, timedelta
from i18n import get_translations, SUPPORTED_LANGS, DEFAULT_LANG
import os
from extensions import limiter, csrf

# Load .env file if present (dev convenience)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

app = Flask(__name__)

# ── Security: require SECRET_KEY in production ────────────────────────────
_secret = os.environ.get('SECRET_KEY', '')
if not _secret:
    import secrets as _secrets
    _secret = _secrets.token_hex(32)
    print('WARNING: SECRET_KEY not set — using a random key (sessions will reset on restart)')
app.config['SECRET_KEY'] = _secret

# Default: siempre apunta a instance/rts_inventory.db relativo al directorio del script
_default_db = 'sqlite:///' + os.path.join(os.path.dirname(os.path.abspath(__file__)), 'instance', 'rts_inventory.db')
_db_url = os.environ.get('DATABASE_URL', _default_db)
# Azure / Heroku usan "postgres://" pero SQLAlchemy requiere "postgresql://"
if _db_url.startswith('postgres://'):
    _db_url = _db_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = _db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# ── Session security ──────────────────────────────────────────────────────
app.config['PERMANENT_SESSION_LIFETIME']  = timedelta(hours=8)
app.config['SESSION_COOKIE_SECURE']       = os.environ.get('FLASK_ENV') == 'production'
app.config['SESSION_COOKIE_HTTPONLY']     = True
app.config['SESSION_COOKIE_SAMESITE']     = 'Lax'

# ── CSRF + Rate limiter (from extensions.py — shared to avoid circular imports) ───
csrf.init_app(app)
limiter.init_app(app)

db.init_app(app)

# ── Búsqueda sin tildes ───────────────────────────────────────────────────────
import unicodedata as _ucd
from sqlalchemy import event as _sa_event
from sqlalchemy.engine import Engine as _Engine

def _nrm(text):
    """Normaliza texto: minúsculas + quita acentos/diacríticos."""
    if text is None:
        return ''
    return ''.join(
        c for c in _ucd.normalize('NFD', str(text).lower())
        if _ucd.category(c) != 'Mn'
    )

@_sa_event.listens_for(_Engine, 'connect')
def _register_sqlite_nrm(dbapi_conn, _rec):
    """Registra nrm() en cada conexión SQLite nueva (no aplica a PostgreSQL)."""
    try:
        dbapi_conn.create_function('nrm', 1, _nrm)
    except Exception:
        pass  # PostgreSQL u otro dialecto — se ignora

def _search_col(col, nq):
    """Expresión LIKE para buscar sin distinguir acentos ni mayúsculas.
    - SQLite: usa la función personalizada nrm() registrada en el engine.
    - PostgreSQL: usa unaccent() + lower() nativos.
    nq debe venir ya normalizado con _nrm()."""
    dialect = db.engine.dialect.name
    if dialect == 'postgresql':
        return db.func.lower(db.func.unaccent(col)).like(f'%{nq}%')
    return db.func.nrm(col).like(f'%{nq}%')

# ── Auth blueprint ────────────────────────────────────────────────────────────
from auth import auth_bp, login_required, module_required, admin_required
app.register_blueprint(auth_bp)

# ── Admin blueprint ───────────────────────────────────────────────────────────
from admin import admin_bp
app.register_blueprint(admin_bp)

# ── Projects blueprint ────────────────────────────────────────────────────────
from projects import projects_bp
app.register_blueprint(projects_bp)

# ── Evaluation blueprint ──────────────────────────────────────────────────────
from eval import eval_bp
app.register_blueprint(eval_bp)

# ── Repository blueprint ──────────────────────────────────────────────────────
from repo import repo_bp
app.register_blueprint(repo_bp)

# ── Reports blueprint ─────────────────────────────────────────────────────────
from reports import reports_bp
app.register_blueprint(reports_bp)

# ── Jinja2 custom filters ─────────────────────────────────────────────────────
import json as _json_mod

@app.template_filter('from_json')
def from_json_filter(value):
    """Deserializa JSON almacenado en la DB para usarlo en templates."""
    if not value:
        return []
    try:
        return _json_mod.loads(value)
    except Exception:
        return []

# ── Security headers ──────────────────────────────────────────────────────
@app.after_request
def set_security_headers(response):
    response.headers['X-Frame-Options']           = 'DENY'
    response.headers['X-Content-Type-Options']    = 'nosniff'
    response.headers['X-XSS-Protection']          = '1; mode=block'
    response.headers['Referrer-Policy']           = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy']        = 'geolocation=(), microphone=(), camera=()'
    if os.environ.get('FLASK_ENV') == 'production':
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    csp = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' cdn.jsdelivr.net; "
        "style-src 'self' 'unsafe-inline' cdn.jsdelivr.net fonts.googleapis.com; "
        "font-src 'self' fonts.gstatic.com cdn.jsdelivr.net; "
        "img-src 'self' data:; "
        "connect-src 'self';"
    )
    response.headers['Content-Security-Policy'] = csp
    return response


# ── Idle timeout (30 min inactivity) ─────────────────────────────────────
IDLE_TIMEOUT = timedelta(minutes=30)

@app.before_request
def check_idle_timeout():
    if not session.get('user'):
        return
    last = session.get('_last_activity')
    now  = datetime.utcnow()
    if last:
        from datetime import datetime as _dt
        last_dt = _dt.fromisoformat(last)
        if now - last_dt > IDLE_TIMEOUT:
            session.clear()
            flash('Tu sesión expiró por inactividad.', 'warning')
            from flask import redirect as _red, url_for as _ufor
            return _red(_ufor('auth.login'))
    session['_last_activity'] = now.isoformat()
    session.modified = True


with app.app_context():
    db.create_all()

    # ── PostgreSQL: habilitar extensión unaccent (búsqueda sin tildes) ────────
    from sqlalchemy import text, inspect as sa_inspect
    if db.engine.dialect.name == 'postgresql':
        try:
            with db.engine.connect() as _conn:
                _conn.execute(text('CREATE EXTENSION IF NOT EXISTS unaccent'))
                _conn.commit()
        except Exception:
            pass

    # ── Migración: agregar columnas nuevas si no existen ─────────────────────
    def _add_col_if_missing(table, col, col_type):
        try:
            insp = sa_inspect(db.engine)
            existing = [c['name'] for c in insp.get_columns(table)]
            if col not in existing:
                with db.engine.connect() as conn:
                    # PostgreSQL soporta IF NOT EXISTS; SQLite lo ignora igual
                    conn.execute(text(f'ALTER TABLE {table} ADD COLUMN IF NOT EXISTS {col} {col_type}'))
                    conn.commit()
        except Exception as _e:
            # SQLite <3.35 no tiene IF NOT EXISTS en ALTER TABLE — fallback
            try:
                insp2 = sa_inspect(db.engine)
                existing2 = [c['name'] for c in insp2.get_columns(table)]
                if col not in existing2:
                    with db.engine.connect() as conn2:
                        conn2.execute(text(f'ALTER TABLE {table} ADD COLUMN {col} {col_type}'))
                        conn2.commit()
            except Exception:
                pass

    for col, typ in [('ram','VARCHAR(50)'), ('os_version','VARCHAR(100)'),
                     ('cpu','VARCHAR(150)'), ('supplier','VARCHAR(150)'),
                     ('last_maintenance','DATE'), ('client_id','INTEGER'),
                     ('department_id', 'INTEGER')]:
        _add_col_if_missing('assets', col, typ)

    # module_access + department_id on users
    _add_col_if_missing('users', 'module_access', 'TEXT')
    _add_col_if_missing('users', 'department_id', 'INTEGER')

    # security fields on users
    for col, typ in [('totp_secret', 'VARCHAR(32)'), ('mfa_enabled', 'INTEGER DEFAULT 0'),
                     ('force_password_change', 'INTEGER DEFAULT 0'),
                     ('failed_logins', 'INTEGER DEFAULT 0'),
                     ('locked_until', 'DATETIME')]:
        _add_col_if_missing('users', col, typ)

    # task_comments and project_activities — created by db.create_all() above
    for col, typ in [('icon', 'VARCHAR(50)'), ('color', 'VARCHAR(20)')]:
        _add_col_if_missing('project_activities', col, typ)

    # access_requests created by db.create_all() above; ensure columns exist
    for col, typ in [('department_id', 'INTEGER'), ('reason', 'TEXT'),
                     ('admin_notes', 'TEXT'), ('reviewed_by', 'VARCHAR(150)'),
                     ('reviewed_at', 'DATETIME')]:
        _add_col_if_missing('access_requests', col, typ)

    # brand_id on assets
    _add_col_if_missing('assets', 'brand_id', 'INTEGER')

    # Extended client fields
    for col, typ in [
        ('location_type', "VARCHAR(10) NOT NULL DEFAULT 'local'"),
        ('contact_name',  'VARCHAR(150)'), ('email',    'VARCHAR(150)'),
        ('phone',         'VARCHAR(50)'),  ('country',  'VARCHAR(80)'),
        ('city',          'VARCHAR(80)'),  ('address',  'VARCHAR(300)'),
        ('rfc',           'VARCHAR(20)'),  ('industry', 'VARCHAR(100)'),
        ('website',       'VARCHAR(200)'), ('start_date','DATE'),
        ('notes',         'TEXT'),
    ]:
        _add_col_if_missing('clients', col, typ)

    # Purchase orders & invoices FK on assets
    _add_col_if_missing('assets', 'purchase_order_id', 'INTEGER')
    _add_col_if_missing('assets', 'invoice_id',        'INTEGER')

    # Maintenance table — created by db.create_all(); ensure all columns exist
    for col, typ in [
        ('ticket_folio',        'VARCHAR(50)'),
        ('maintenance_type',    'VARCHAR(20)'),
        ('status',              'VARCHAR(20)'),
        ('prev_asset_status',   'VARCHAR(30)'),
        ('reported_date',       'DATE'),
        ('reported_by',         'VARCHAR(150)'),
        ('process_name',        'VARCHAR(150)'),
        ('process_responsible', 'VARCHAR(150)'),
        ('nc_source',           'VARCHAR(100)'),
        ('description',         'TEXT'),
        ('analysis_method',     'VARCHAR(150)'),
        ('participants',        'VARCHAR(300)'),
        ('root_cause_analysis', 'TEXT'),
        ('root_cause',          'TEXT'),
        ('correction_desc',     'TEXT'),
        ('action_plan',         'TEXT'),
        ('proposed_close_date', 'DATE'),
        ('followup_responsible','VARCHAR(150)'),
        ('close_responsible',   'VARCHAR(150)'),
        ('effectiveness_ok',    'INTEGER'),
        ('effectiveness_notes', 'TEXT'),
        ('actual_close_date',   'DATE'),
        ('document_path',       'VARCHAR(500)'),
        ('document_name',       'VARCHAR(200)'),
        ('photos',              'TEXT'),
        ('notes',               'TEXT'),
        ('created_at',          'DATETIME'),
        ('updated_at',          'DATETIME'),
    ]:
        _add_col_if_missing('maintenance', col, typ)

    # Licenses tables — created by db.create_all(); ensure all columns exist
    for col, typ in [
        ('name',            'VARCHAR(200)'),
        ('vendor',          'VARCHAR(100)'),
        ('software',        'VARCHAR(150)'),
        ('category',        'VARCHAR(50)'),
        ('license_type',    'VARCHAR(30)'),
        ('license_key',     'VARCHAR(500)'),
        ('is_microsoft',    'INTEGER DEFAULT 0'),
        ('tenant_id',       'VARCHAR(100)'),
        ('tenant_name',     'VARCHAR(200)'),
        ('tenant_domain',   'VARCHAR(200)'),
        ('subscription_id', 'VARCHAR(100)'),
        ('sku_name',        'VARCHAR(100)'),
        ('seat_count',      'INTEGER'),
        ('purchase_cost',   'REAL'),
        ('renewal_cost',    'REAL'),
        ('currency',        "VARCHAR(10) DEFAULT 'MXN'"),
        ('purchase_date',   'DATE'),
        ('expiry_date',     'DATE'),
        ('renewal_date',    'DATE'),
        ('status',          "VARCHAR(20) DEFAULT 'active'"),
        ('notes',           'TEXT'),
        ('created_at',      'DATETIME'),
        ('updated_at',      'DATETIME'),
    ]:
        _add_col_if_missing('licenses', col, typ)

    for col, typ in [
        ('license_id',    'INTEGER'),
        ('employee_id',   'INTEGER'),
        ('asset_id',      'INTEGER'),
        ('assigned_date', 'DATE'),
        ('notes',         'TEXT'),
        ('created_at',    'DATETIME'),
    ]:
        _add_col_if_missing('license_assignments', col, typ)

    # Absolute fields on assets
    for col, typ in [
        ('absolute_id',        'VARCHAR(100)'),
        ('absolute_status',    'VARCHAR(30)'),
        ('absolute_username',  'VARCHAR(150)'),
        ('absolute_last_seen', 'DATETIME'),
        ('absolute_sync_at',   'DATETIME'),
    ]:
        _add_col_if_missing('assets', col, typ)

    # Evaluation tables — created by db.create_all()
    for col, typ in [
        ('empresa',               'VARCHAR(150)'),
        ('localidad',             'VARCHAR(150)'),
        ('nivel',                 'VARCHAR(100)'),
        ('employee_submitted_at', 'DATETIME'),
        ('chief_submitted_at',    'DATETIME'),
        ('updated_at',            'DATETIME'),
    ]:
        _add_col_if_missing('evaluations', col, typ)

    # Employee extended fields
    for col, typ in [
        ('position',   'VARCHAR(150)'),
        ('client_id',  'INTEGER'),
        ('site_type',  "VARCHAR(10) DEFAULT 'sitio'"),
        ('address',    'TEXT'),
        ('whatsapp',   'VARCHAR(50)'),
    ]:
        _add_col_if_missing('employees', col, typ)

    # Seed IT department if none exist
    from models import Department
    if Department.query.count() == 0:
        it = Department(name='IT', code='IT', color='#089ACF',
                        manager_name='IT Department')
        db.session.add(it)
        db.session.commit()
        print('✅  Departamento IT creado')

    # Crear admin por defecto si no existe ningún usuario
    from models import User
    if User.query.count() == 0:
        import secrets as _sec
        tmp_pwd = _sec.token_urlsafe(16)
        admin = User(name='Administrador', username='admin', role='admin',
                     force_password_change=True)
        admin.set_password(tmp_pwd)
        db.session.add(admin)
        db.session.commit()
        print(f'Usuario admin creado  ->  admin / {tmp_pwd}')
        print('Cambia la contrasena en el primer inicio de sesion.')

# ── Static image helpers ──────────────────────────────────────────────────────
STATIC = os.path.join(os.path.dirname(__file__), 'static', 'images')

def _img_exists(filename):
    return os.path.exists(os.path.join(STATIC, filename))


# ── Helpers ───────────────────────────────────────────────────────────────────

def parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return None


# ── Portal ────────────────────────────────────────────────────────────────────

_MODULE_URLS = {
    'inventory':  lambda: url_for('inventory_dashboard'),
    'projects':   lambda: url_for('projects.dashboard'),
    'evaluation': lambda: url_for('eval.index'),
    'repository': lambda: url_for('repo.index'),
}

@app.route('/')
@login_required
def portal():
    user     = session.get('user', {})
    modules  = user.get('modules', [])
    is_admin = user.get('role') == 'admin'
    uid      = user.get('id')

    # Pending access-request slugs for this user
    pending_slugs = set()
    denied_slugs  = set()
    if uid:
        my_reqs = AccessRequest.query.filter_by(user_id=uid).all()
        for r in my_reqs:
            if r.status == 'pending':
                pending_slugs.add(r.module_slug)
            elif r.status == 'denied':
                denied_slugs.add(r.module_slug)

    # Build card list — ALL modules, each tagged with access state
    all_cards = []
    for m in ALL_MODULES:
        card = dict(m)
        slug = m['slug']
        if is_admin or slug in modules:
            card['state'] = 'open'
            card['url']   = _MODULE_URLS[slug]()
        elif slug in pending_slugs:
            card['state'] = 'pending'
        else:
            card['state'] = 'locked'
        all_cards.append(card)

    return render_template('portal.html', all_cards=all_cards,
                           departments=Department.query.filter_by(active=True).all())


# ── Request Access ────────────────────────────────────────────────────────────

@app.route('/request-access/<slug>', methods=['POST'])
@login_required
def request_access(slug):
    user = session.get('user', {})
    uid  = user.get('id')
    # Verify slug is valid
    valid_slugs = [m['slug'] for m in ALL_MODULES]
    if slug not in valid_slugs:
        flash('Invalid module.', 'danger')
        return redirect(url_for('portal'))

    # Don't allow duplicate pending requests
    existing = AccessRequest.query.filter_by(
        user_id=uid, module_slug=slug, status='pending').first()
    if existing:
        flash('You already have a pending request for this module.', 'info')
        return redirect(url_for('portal'))

    dept_id = request.form.get('department_id', type=int) or None
    reason  = request.form.get('reason', '').strip()

    req = AccessRequest(
        user_id=uid,
        user_name=user.get('name'),
        module_slug=slug,
        department_id=dept_id,
        reason=reason,
        status='pending',
    )
    db.session.add(req)
    log_action('create', 'access_request',
               entity_name=f'{user.get("name")} → {slug}',
               details=reason[:200] if reason else None)
    db.session.commit()

    # Notify admins via Teams
    import notifications as notif
    cfg = notif.load_config()
    if cfg.get('enabled') and cfg.get('teams_enabled'):
        mod_name = next((m['name'] for m in ALL_MODULES if m['slug'] == slug), slug)
        notif.send_teams(
            f'Access Request: {mod_name}',
            f'{user.get("name")} is requesting access to **{mod_name}**.',
            facts=[('User', user.get('name')), ('Module', mod_name),
                   ('Reason', reason[:120] if reason else 'No reason given')],
            url=notif._url(cfg, '/admin/access-requests'),
            color='FFA000'
        )

    flash('Your request has been sent to IT. You\'ll be notified once it\'s reviewed.', 'success')
    return redirect(url_for('portal'))


# ── Inventory Dashboard ───────────────────────────────────────────────────────

@app.route('/inventory')
@module_required('inventory')
def inventory_dashboard():
    from models import User as _User
    sess_user   = session.get('user', {})
    is_admin    = sess_user.get('role') == 'admin'
    db_user     = _User.query.get(sess_user.get('id')) if sess_user.get('id') else None
    # Department-scoped access: non-admin users with a department see only that dept's assets
    dept_filter = None
    if not is_admin and db_user and db_user.department_id:
        dept_filter = db_user.department_id

    def _asset_q():
        q = Asset.query
        if dept_filter:
            q = q.filter_by(department_id=dept_filter)
        return q

    total_assets      = _asset_q().count()
    available         = _asset_q().filter_by(status='available').count()
    in_use            = _asset_q().filter_by(status='in_use').count()
    maintenance       = _asset_q().filter_by(status='maintenance').count()
    retired           = _asset_q().filter_by(status='retired').count()
    total_employees   = Employee.query.filter_by(active=True).count()
    total_categories  = Category.query.count()
    foraneo_count     = Asset.query.filter_by(location_type='foraneo').count()
    in_transit_count  = Shipment.query.filter_by(status='en_transito').count()

    # Empleados activos por cliente (para el panel "Personas por Cliente")
    client_emp_rows = db.session.query(
        Client.id, Client.name,
        db.func.count(Employee.id).label('emp_count')
    ).outerjoin(Employee, db.and_(
        Employee.client_id == Client.id,
        Employee.active == True        # noqa: E712
    )).group_by(Client.id, Client.name
    ).order_by(db.desc('emp_count'), Client.name).all()
    # Empleados sin cliente asignado
    unassigned_emp = Employee.query.filter_by(active=True, client_id=None).count()
    client_employee_counts = [
        {'id': r.id, 'name': r.name, 'count': r.emp_count}
        for r in client_emp_rows
    ]
    if unassigned_emp:
        client_employee_counts.append({'id': None, 'name': 'Sin cliente', 'count': unassigned_emp})
    recent_assignments = Assignment.query.order_by(Assignment.created_at.desc()).limit(8).all()
    recent_shipments  = Shipment.query.filter(
        Shipment.status.notin_(['entregado', 'devuelto'])
    ).order_by(Shipment.created_at.desc()).limit(5).all()
    categories  = Category.query.all()
    cat_stats   = [{'name': c.name, 'count': len(c.assets)} for c in categories]

    # ── Gráfica: datos raw (dicts por mes) para todas las fechas ─────────────
    # 4 queries totales — el JS resuelve cualquier rango/categoría client-side

    cats_for_chart = Category.query.order_by(Category.name).all()

    # 1) Adquisiciones por (category_id, YYYY-MM)
    acq_raw = db.session.query(
        Asset.category_id,
        db.func.strftime('%Y-%m', Asset.created_at),
        db.func.count(Asset.id)
    ).group_by(Asset.category_id,
               db.func.strftime('%Y-%m', Asset.created_at)).all()

    # 2) Bajas globales del audit log por YYYY-MM
    baja_raw = db.session.query(
        db.func.strftime('%Y-%m', AuditLog.created_at),
        db.func.count(AuditLog.id)
    ).filter(
        AuditLog.entity_type == 'asset',
        db.or_(
            AuditLog.action == 'delete',
            db.and_(AuditLog.action == 'update',
                    db.or_(AuditLog.details.ilike('%→ retired%'),
                           AuditLog.details.ilike('%→ disposed%')))
        )
    ).group_by(db.func.strftime('%Y-%m', AuditLog.created_at)).all()

    # 3) Bajas por categoría (activos retired/disposed, fecha updated_at)
    baja_cat_raw = db.session.query(
        Asset.category_id,
        db.func.strftime('%Y-%m', Asset.updated_at),
        db.func.count(Asset.id)
    ).filter(Asset.status.in_(['retired', 'disposed'])
    ).group_by(Asset.category_id,
               db.func.strftime('%Y-%m', Asset.updated_at)).all()

    # 4) Reasignaciones por (category_id, YYYY-MM)
    reas_raw = db.session.query(
        Asset.category_id,
        db.func.strftime('%Y-%m', Assignment.created_at),
        db.func.count(Assignment.id)
    ).join(Asset, Assignment.asset_id == Asset.id
    ).group_by(Asset.category_id,
               db.func.strftime('%Y-%m', Assignment.created_at)).all()

    # ── Construir estructura de dicts raw por categoría ───────────────────────
    cat_chart_raw = {
        'all': {'adquisiciones': {}, 'bajas': {}, 'reasignaciones': {}}
    }
    for cat in cats_for_chart:
        cat_chart_raw[str(cat.id)] = {'adquisiciones': {}, 'bajas': {}, 'reasignaciones': {}}

    for cat_id, ym, cnt in acq_raw:
        if not ym:
            continue
        cat_chart_raw['all']['adquisiciones'][ym] = \
            cat_chart_raw['all']['adquisiciones'].get(ym, 0) + cnt
        if cat_id and str(cat_id) in cat_chart_raw:
            cat_chart_raw[str(cat_id)]['adquisiciones'][ym] = cnt

    for ym, cnt in baja_raw:
        if ym:
            cat_chart_raw['all']['bajas'][ym] = cnt

    for cat_id, ym, cnt in baja_cat_raw:
        if ym and cat_id and str(cat_id) in cat_chart_raw:
            cat_chart_raw[str(cat_id)]['bajas'][ym] = cnt

    for cat_id, ym, cnt in reas_raw:
        if not ym:
            continue
        cat_chart_raw['all']['reasignaciones'][ym] = \
            cat_chart_raw['all']['reasignaciones'].get(ym, 0) + cnt
        if cat_id and str(cat_id) in cat_chart_raw:
            cat_chart_raw[str(cat_id)]['reasignaciones'][ym] = cnt

    chart_start_year = 2020

    active_dept = Department.query.get(dept_filter) if dept_filter else None

    # Warranty expiry alerts (30 / 60 days)
    soon_30 = date.today() + timedelta(days=30)
    soon_60 = date.today() + timedelta(days=60)
    warranty_expiring = _asset_q().filter(
        Asset.warranty_expiry != None,          # noqa: E711
        Asset.warranty_expiry <= soon_60,
        Asset.warranty_expiry >= date.today(),
        Asset.status.notin_(['retired', 'disposed'])
    ).order_by(Asset.warranty_expiry).limit(10).all()

    return render_template('inventory/dashboard.html',
                           total_assets=total_assets, available=available,
                           in_use=in_use, maintenance=maintenance, retired=retired,
                           total_employees=total_employees, total_categories=total_categories,
                           foraneo_count=foraneo_count, in_transit_count=in_transit_count,
                           recent_assignments=recent_assignments,
                           recent_shipments=recent_shipments,
                           cat_stats=cat_stats,
                           cat_chart_raw=cat_chart_raw,
                           cats_for_chart=cats_for_chart,
                           chart_start_year=chart_start_year,
                           chart_current_year=date.today().year,
                           active_dept=active_dept,
                           warranty_expiring=warranty_expiring,
                           soon_30=soon_30,
                           client_employee_counts=client_employee_counts,
                           departments=Department.query.filter_by(active=True).order_by(Department.name).all())


# ── Assets ────────────────────────────────────────────────────────────────────

@app.route('/assets')
@login_required
def assets_list():
    q               = request.args.get('q', '')
    status_filter   = request.args.get('status', '')
    category_filter = request.args.get('category', '')
    loc_filter      = request.args.get('location_type', '')
    query = Asset.query
    if q:
        nq = _nrm(q)
        # Buscar también por nombre de empleado asignado
        emp_asset_ids = db.session.query(Assignment.asset_id).join(
            Employee, Assignment.employee_id == Employee.id
        ).filter(
            Assignment.returned_date == None,   # noqa: E711
            _search_col(Employee.name, nq)
        ).subquery()
        query = query.outerjoin(Client, Asset.client_id == Client.id).filter(db.or_(
            _search_col(Asset.name, nq),        _search_col(Asset.asset_tag, nq),
            _search_col(Asset.serial_number, nq), _search_col(Asset.manufacturer, nq),
            _search_col(Asset.model, nq),       _search_col(Asset.supplier, nq),
            _search_col(Asset.location, nq),    _search_col(Asset.cpu, nq),
            _search_col(Asset.ram, nq),         _search_col(Client.name, nq),
            Asset.id.in_(emp_asset_ids),
        ))
    if status_filter:
        query = query.filter_by(status=status_filter)
    if category_filter:
        query = query.filter_by(category_id=category_filter)
    if loc_filter:
        query = query.filter_by(location_type=loc_filter)
    page       = request.args.get('page', 1, type=int)
    per_page   = 50
    pagination = query.order_by(Asset.asset_tag).paginate(page=page, per_page=per_page, error_out=False)
    assets     = pagination.items
    categories = Category.query.order_by(Category.name).all()
    return render_template('assets/list.html', assets=assets, categories=categories,
                           q=q, status_filter=status_filter,
                           category_filter=category_filter, loc_filter=loc_filter,
                           pagination=pagination)


@app.route('/api/assets/autocomplete')
@login_required
def assets_autocomplete():
    """Devuelve sugerencias JSON para el buscador de activos."""
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify([])
    nq = _nrm(q)

    # Búsqueda directa por campos del activo
    direct = Asset.query.outerjoin(Client, Asset.client_id == Client.id).filter(db.or_(
        _search_col(Asset.name, nq),        _search_col(Asset.asset_tag, nq),
        _search_col(Asset.serial_number, nq), _search_col(Asset.manufacturer, nq),
        _search_col(Asset.model, nq),       _search_col(Asset.supplier, nq),
        _search_col(Client.name, nq),
    )).limit(10).all()

    # Búsqueda por empleado asignado
    by_emp = Asset.query.join(Assignment, Asset.id == Assignment.asset_id).join(
        Employee, Assignment.employee_id == Employee.id
    ).filter(
        Assignment.returned_date == None,   # noqa: E711
        _search_col(Employee.name, nq)
    ).limit(6).all()

    seen = set()
    results = []
    for a in direct + by_emp:
        if a.id in seen:
            continue
        seen.add(a.id)
        # Empleado asignado actualmente
        asn = Assignment.query.filter_by(asset_id=a.id, returned_date=None).first()
        emp = asn.employee.name if asn and asn.employee else None
        results.append({
            'id':       a.id,
            'name':     a.name,
            'tag':      a.asset_tag or '',
            'serial':   a.serial_number or '',
            'employee': emp or '',
            'status':   a.status,
            'url':      url_for('asset_detail', id=a.id),
        })

    return jsonify(results[:12])


@app.route('/assets/new', methods=['GET', 'POST'])
@login_required
def asset_new():
    categories      = Category.query.order_by(Category.name).all()
    clients         = Client.query.filter_by(active=True).order_by(Client.name).all()
    brands          = Brand.query.filter_by(active=True).order_by(Brand.name).all()
    purchase_orders = PurchaseOrder.query.order_by(PurchaseOrder.created_at.desc()).all()
    invoices        = Invoice.query.order_by(Invoice.created_at.desc()).all()
    return_url      = request.args.get('return_url') or url_for('assets_list')
    if request.method == 'POST':
        return_url = request.form.get('return_url') or return_url
        asset_tag  = request.form.get('asset_tag', '').strip()
        if Asset.query.filter_by(asset_tag=asset_tag).first():
            flash(f'El Asset Tag "{asset_tag}" ya existe.', 'danger')
            return render_template('assets/form.html', categories=categories,
                                   clients=clients, brands=brands,
                                   purchase_orders=purchase_orders, invoices=invoices,
                                   asset=None, asset_type_choices=Asset.ASSET_TYPE_CHOICES,
                                   form=request.form, return_url=return_url)
        asset = Asset(
            name=request.form.get('name', '').strip(),
            asset_tag=asset_tag,
            asset_type=request.form.get('asset_type', 'laptop'),
            serial_number=request.form.get('serial_number', '').strip() or None,
            manufacturer=request.form.get('manufacturer', '').strip() or None,
            model=request.form.get('model', '').strip() or None,
            ram=request.form.get('ram', '').strip() or None,
            cpu=request.form.get('cpu', '').strip() or None,
            os_version=request.form.get('os_version', '').strip() or None,
            category_id=request.form.get('category_id') or None,
            client_id=request.form.get('client_id') or None,
            brand_id=request.form.get('brand_id') or None,
            purchase_order_id=request.form.get('purchase_order_id') or None,
            invoice_id=request.form.get('invoice_id') or None,
            status=request.form.get('status', 'available'),
            location_type=request.form.get('location_type', 'en_sitio'),
            location=request.form.get('location', '').strip() or None,
            purchase_date=parse_date(request.form.get('purchase_date')),
            purchase_cost=float(request.form.get('purchase_cost')) if request.form.get('purchase_cost') else None,
            supplier=request.form.get('supplier', '').strip() or None,
            warranty_expiry=parse_date(request.form.get('warranty_expiry')),
            last_maintenance=parse_date(request.form.get('last_maintenance')),
            notes=request.form.get('notes', '').strip() or None,
        )
        db.session.add(asset)
        log_action('create', 'asset', entity_name=asset.name,
                   details=f'Tag: {asset_tag} | Estado: {asset.status} | Tipo: {asset.location_type}')
        db.session.commit()
        flash(f'Activo "{asset.name}" creado correctamente.', 'success')
        return redirect(return_url)
    return render_template('assets/form.html', categories=categories,
                           clients=clients, brands=brands,
                           purchase_orders=purchase_orders, invoices=invoices,
                           asset=None, asset_type_choices=Asset.ASSET_TYPE_CHOICES,
                           form={}, return_url=return_url)


@app.route('/assets/<int:id>')
@login_required
def asset_detail(id):
    asset      = Asset.query.get_or_404(id)
    return_url = request.args.get('return_url') or url_for('assets_list')
    # Per-asset audit history
    history = AuditLog.query.filter(
        AuditLog.entity_type == 'asset',
        AuditLog.entity_id   == id,
    ).order_by(AuditLog.created_at.desc()).limit(50).all()
    abs_configured = bool(AppSetting.get('absolute_token_id'))
    return render_template('assets/detail.html', asset=asset, return_url=return_url,
                           history=history, absolute_configured=abs_configured)


@app.route('/assets/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def asset_edit(id):
    asset      = Asset.query.get_or_404(id)
    categories      = Category.query.order_by(Category.name).all()
    clients         = Client.query.filter_by(active=True).order_by(Client.name).all()
    brands          = Brand.query.filter_by(active=True).order_by(Brand.name).all()
    purchase_orders = PurchaseOrder.query.order_by(PurchaseOrder.created_at.desc()).all()
    invoices        = Invoice.query.order_by(Invoice.created_at.desc()).all()
    return_url      = request.args.get('return_url') or url_for('assets_list')
    if request.method == 'POST':
        return_url = request.form.get('return_url') or return_url
        new_tag    = request.form.get('asset_tag', '').strip()
        existing   = Asset.query.filter_by(asset_tag=new_tag).first()
        if existing and existing.id != asset.id:
            flash(f'El Asset Tag "{new_tag}" ya existe en otro activo.', 'danger')
            return render_template('assets/form.html', categories=categories,
                                   clients=clients, brands=brands,
                                   purchase_orders=purchase_orders, invoices=invoices,
                                   asset=asset, asset_type_choices=Asset.ASSET_TYPE_CHOICES,
                                   form=request.form, return_url=return_url)
        changes = []
        if asset.name != request.form.get('name', '').strip():
            changes.append(f'nombre: {asset.name} → {request.form.get("name").strip()}')
        if asset.status != request.form.get('status'):
            changes.append(f'estado: {asset.status} → {request.form.get("status")}')
        if asset.location_type != request.form.get('location_type'):
            changes.append(f'tipo: {asset.location_type} → {request.form.get("location_type")}')
        new_client_id = request.form.get('client_id') or None
        if str(asset.client_id or '') != str(new_client_id or ''):
            changes.append('cliente actualizado')

        asset.name              = request.form.get('name', '').strip()
        asset.asset_tag         = new_tag
        asset.asset_type        = request.form.get('asset_type', asset.asset_type or 'laptop')
        asset.serial_number     = request.form.get('serial_number', '').strip() or None
        asset.manufacturer      = request.form.get('manufacturer', '').strip() or None
        asset.model             = request.form.get('model', '').strip() or None
        asset.ram               = request.form.get('ram', '').strip() or None
        asset.cpu               = request.form.get('cpu', '').strip() or None
        asset.os_version        = request.form.get('os_version', '').strip() or None
        asset.category_id       = request.form.get('category_id') or None
        asset.client_id         = new_client_id
        asset.brand_id          = request.form.get('brand_id') or None
        asset.purchase_order_id = request.form.get('purchase_order_id') or None
        asset.invoice_id        = request.form.get('invoice_id') or None
        asset.status            = request.form.get('status', 'available')
        asset.location_type     = request.form.get('location_type', 'en_sitio')
        asset.location          = request.form.get('location', '').strip() or None
        asset.purchase_date     = parse_date(request.form.get('purchase_date'))
        asset.purchase_cost     = float(request.form.get('purchase_cost')) if request.form.get('purchase_cost') else None
        asset.supplier          = request.form.get('supplier', '').strip() or None
        asset.warranty_expiry   = parse_date(request.form.get('warranty_expiry'))
        asset.last_maintenance  = parse_date(request.form.get('last_maintenance'))
        asset.notes             = request.form.get('notes', '').strip() or None
        log_action('update', 'asset', entity_id=asset.id, entity_name=asset.name,
                   details='; '.join(changes) if changes else 'Actualización sin cambios clave')
        db.session.commit()
        flash(f'Activo "{asset.name}" actualizado.', 'success')
        return redirect(return_url)
    return render_template('assets/form.html', categories=categories,
                           clients=clients, brands=brands,
                           purchase_orders=purchase_orders, invoices=invoices,
                           asset=asset, asset_type_choices=Asset.ASSET_TYPE_CHOICES,
                           form={}, return_url=return_url)


@app.route('/assets/<int:id>/delete', methods=['POST'])
@login_required
def asset_delete(id):
    asset = Asset.query.get_or_404(id)
    name  = asset.name
    tag   = asset.asset_tag
    log_action('delete', 'asset', entity_id=asset.id, entity_name=name,
               details=f'Activo eliminado: {tag}')
    db.session.delete(asset)
    db.session.commit()
    flash(f'Activo "{name}" eliminado.', 'warning')
    return redirect(url_for('assets_list'))


# ── Employees ─────────────────────────────────────────────────────────────────

@app.route('/employees')
@login_required
def employees_list():
    q             = request.args.get('q', '')
    show_inactive = request.args.get('inactive', '')
    fclient       = request.args.get('client_id', '', type=str)
    fsite         = request.args.get('site_type', '')
    query = Employee.query
    if not show_inactive:
        query = query.filter_by(active=True)
    if fclient:
        query = query.filter_by(client_id=int(fclient))
    if fsite:
        query = query.filter_by(site_type=fsite)
    if q:
        nq = _nrm(q)
        query = query.filter(db.or_(
            _search_col(Employee.name, nq),        _search_col(Employee.employee_id, nq),
            _search_col(Employee.department, nq),  _search_col(Employee.email, nq),
        ))
    page       = request.args.get('page', 1, type=int)
    pagination = query.order_by(Employee.name).paginate(page=page, per_page=50, error_out=False)
    employees  = pagination.items
    clients    = Client.query.order_by(Client.name).all()
    return render_template('employees/list.html', employees=employees,
                           q=q, show_inactive=show_inactive, pagination=pagination,
                           clients=clients, fclient=fclient, fsite=fsite)


@app.route('/employees/new', methods=['GET', 'POST'])
@login_required
def employee_new():
    clients = Client.query.order_by(Client.name).all()
    if request.method == 'POST':
        emp_id = request.form.get('employee_id', '').strip()
        if Employee.query.filter_by(employee_id=emp_id).first():
            flash(f'El ID de empleado "{emp_id}" ya existe.', 'danger')
            return render_template('employees/form.html', employee=None,
                                   form=request.form, clients=clients)
        client_id = request.form.get('client_id') or None
        emp = Employee(
            name=request.form.get('name', '').strip(),
            employee_id=emp_id,
            position=request.form.get('position', '').strip() or None,
            department=request.form.get('department', '').strip() or None,
            email=request.form.get('email', '').strip() or None,
            phone=request.form.get('phone', '').strip() or None,
            whatsapp=request.form.get('whatsapp', '').strip() or None,
            client_id=int(client_id) if client_id else None,
            site_type=request.form.get('site_type', 'sitio'),
            address=request.form.get('address', '').strip() or None,
            active=True,
        )
        db.session.add(emp)
        log_action('create', 'employee', entity_name=emp.name,
                   details=f'ID: {emp_id} | Depto: {emp.department}')
        db.session.commit()
        flash(f'Empleado "{emp.name}" creado correctamente.', 'success')
        return redirect(url_for('employees_list'))
    return render_template('employees/form.html', employee=None, form={}, clients=clients)


@app.route('/employees/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def employee_edit(id):
    emp     = Employee.query.get_or_404(id)
    clients = Client.query.order_by(Client.name).all()
    if request.method == 'POST':
        new_emp_id = request.form.get('employee_id', '').strip()
        existing   = Employee.query.filter_by(employee_id=new_emp_id).first()
        if existing and existing.id != emp.id:
            flash(f'El ID "{new_emp_id}" ya existe en otro empleado.', 'danger')
            return render_template('employees/form.html', employee=emp,
                                   form=request.form, clients=clients)
        changes = []
        if emp.name != request.form.get('name', '').strip():
            changes.append(f'nombre: {emp.name} → {request.form.get("name").strip()}')
        if emp.department != (request.form.get('department', '').strip() or None):
            changes.append(f'depto: {emp.department} → {request.form.get("department")}')
        client_id = request.form.get('client_id') or None
        emp.name        = request.form.get('name', '').strip()
        emp.employee_id = new_emp_id
        emp.position    = request.form.get('position', '').strip() or None
        emp.department  = request.form.get('department', '').strip() or None
        emp.email       = request.form.get('email', '').strip() or None
        emp.phone       = request.form.get('phone', '').strip() or None
        emp.whatsapp    = request.form.get('whatsapp', '').strip() or None
        emp.client_id   = int(client_id) if client_id else None
        emp.site_type   = request.form.get('site_type', 'sitio')
        emp.address     = request.form.get('address', '').strip() or None
        emp.active      = 'active' in request.form
        log_action('update', 'employee', entity_id=emp.id, entity_name=emp.name,
                   details='; '.join(changes) if changes else 'Sin cambios clave')
        db.session.commit()
        flash(f'Empleado "{emp.name}" actualizado.', 'success')
        return redirect(url_for('employees_list'))
    return render_template('employees/form.html', employee=emp, form={}, clients=clients)


@app.route('/employees/<int:id>/delete', methods=['POST'])
@login_required
def employee_delete(id):
    emp  = Employee.query.get_or_404(id)
    name = emp.name
    log_action('delete', 'employee', entity_id=emp.id, entity_name=name,
               details=f'Empleado eliminado: {emp.employee_id}')
    db.session.delete(emp)
    db.session.commit()
    flash(f'Empleado "{name}" eliminado.', 'warning')
    return redirect(url_for('employees_list'))


@app.route('/employees/<int:id>/responsiva')
@login_required
def employee_responsiva(id):
    """Show asset selection page for generating Carta de Resguardo."""
    emp = Employee.query.get_or_404(id)
    # Get current active assignments (not returned)
    active_assignments = Assignment.query.filter_by(
        employee_id=emp.id, returned_date=None
    ).all()
    assets = [a.asset for a in active_assignments]
    return render_template('employees/responsiva.html', employee=emp,
                           assets=assets)


@app.route('/employees/<int:id>/responsiva/download', methods=['POST'])
@login_required
def employee_responsiva_download(id):
    """Generate and return Carta de Resguardo as non-editable PDF."""
    from responsiva_pdf import generate_responsiva_pdf
    emp = Employee.query.get_or_404(id)

    # Selected asset IDs from checkboxes
    selected_ids = request.form.getlist('asset_ids', type=int)
    if not selected_ids:
        flash('Selecciona al menos un activo para generar la carta.', 'warning')
        return redirect(url_for('employee_responsiva', id=id))

    assets = Asset.query.filter(Asset.id.in_(selected_ids)).all()
    if not assets:
        flash('No se encontraron los activos seleccionados.', 'danger')
        return redirect(url_for('employee_responsiva', id=id))

    # Use assignment date of earliest selected assignment, or today
    earliest = Assignment.query.filter(
        Assignment.employee_id == emp.id,
        Assignment.asset_id.in_(selected_ids),
        Assignment.returned_date == None
    ).order_by(Assignment.assigned_date.asc()).first()
    assign_date = earliest.assigned_date if earliest else date.today()

    buf = generate_responsiva_pdf(emp, assets, assign_date=assign_date)

    safe_name = emp.name.replace(' ', '_')
    filename = f'Responsiva_{safe_name}_{assign_date.strftime("%Y%m%d")}.pdf'

    log_action('export', 'employee', entity_id=emp.id, entity_name=emp.name,
               details=f'Carta de Resguardo PDF generada ({len(assets)} activos)')

    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )


# ── Offboarding ───────────────────────────────────────────────────────────────

@app.route('/employees/<int:id>/offboarding')
@login_required
def employee_offboarding(id):
    """Pantalla de offboarding: muestra activos con depreciación y estado."""
    from offboarding_pdf import calc_depreciation, OFFBOARDING_REASONS
    emp = Employee.query.get_or_404(id)
    active_assignments = Assignment.query.filter_by(
        employee_id=emp.id, returned_date=None
    ).all()
    assets = [a.asset for a in active_assignments]

    # Pre-calcula depreciación para cada activo (para mostrar en UI)
    asset_deprs = {}
    for a in assets:
        asset_deprs[a.id] = calc_depreciation(
            a.purchase_cost, a.purchase_date, a.asset_type or 'otro'
        )

    return render_template('employees/offboarding.html',
                           employee=emp, assets=assets,
                           asset_deprs=asset_deprs,
                           reasons=OFFBOARDING_REASONS)


@app.route('/employees/<int:id>/offboarding/pdf', methods=['POST'])
@login_required
def employee_offboarding_pdf(id):
    """Genera el PDF de Acta de Entrega-Recepción."""
    from offboarding_pdf import (generate_offboarding_pdf,
                                 calc_depreciation, OFFBOARDING_REASONS)
    emp = Employee.query.get_or_404(id)

    selected_ids  = request.form.getlist('asset_ids', type=int)
    reason        = request.form.get('reason', 'otro')
    off_date_raw  = request.form.get('offboarding_date', '')
    off_date      = _parse_date(off_date_raw) or date.today()
    return_assets = 'return_assets' in request.form   # checkbox: marcar como devueltos

    if not selected_ids:
        flash('Selecciona al menos un activo.', 'warning')
        return redirect(url_for('employee_offboarding', id=id))

    assets = Asset.query.filter(Asset.id.in_(selected_ids)).all()

    # Armar entries con condición y daño
    asset_entries = []
    for a in assets:
        cond   = request.form.get(f'cond_{a.id}', 'bueno')
        dnotes = request.form.get(f'dmg_{a.id}', '').strip()
        depr   = calc_depreciation(a.purchase_cost, a.purchase_date, a.asset_type or 'otro',
                                   as_of=off_date)
        asset_entries.append({
            'asset':        a,
            'condition':    cond,
            'damage_notes': dnotes,
            'depr':         depr,
        })

    # Marcar activos como devueltos si se marcó el checkbox
    if return_assets:
        for a in assets:
            asn = Assignment.query.filter_by(
                asset_id=a.id, employee_id=emp.id, returned_date=None
            ).first()
            if asn:
                asn.returned_date = off_date
                a.status = asn.asset.status if asn.asset.status != 'in_use' else 'available'
                a.status = 'available'
        db.session.commit()

    buf = generate_offboarding_pdf(emp, asset_entries,
                                   offboarding_date=off_date, reason=reason)

    safe_name = emp.name.replace(' ', '_')
    filename  = f'Offboarding_{safe_name}_{off_date.strftime("%Y%m%d")}.pdf'

    log_action('export', 'employee', entity_id=emp.id, entity_name=emp.name,
               details=f'Acta Offboarding PDF ({len(assets)} activos, devueltos={return_assets})')

    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/pdf')


# ── Assignments ───────────────────────────────────────────────────────────────

@app.route('/assignments')
@login_required
def assignments_list():
    active_only = request.args.get('active', '1')
    query = Assignment.query
    if active_only == '1':
        query = query.filter_by(returned_date=None)
    assignments = query.order_by(Assignment.assigned_date.desc()).all()
    return render_template('assignments/list.html',
                           assignments=assignments, active_only=active_only)


@app.route('/assignments/new', methods=['GET', 'POST'])
@login_required
def assignment_new():
    assets    = Asset.query.filter_by(status='available').order_by(Asset.name).all()
    employees = Employee.query.filter_by(active=True).order_by(Employee.name).all()
    pre_asset = request.args.get('asset_id')
    if request.method == 'POST':
        asset_id    = int(request.form.get('asset_id'))
        employee_id = int(request.form.get('employee_id'))
        asset = Asset.query.get_or_404(asset_id)
        if asset.current_assignment:
            flash('Este activo ya está asignado. Devuélvelo primero.', 'danger')
            return redirect(url_for('assignments_list'))
        emp = Employee.query.get_or_404(employee_id)
        assignment = Assignment(
            asset_id=asset_id,
            employee_id=employee_id,
            assigned_date=parse_date(request.form.get('assigned_date')) or date.today(),
            notes=request.form.get('notes', '').strip() or None,
        )
        asset.status = 'in_use'
        db.session.add(assignment)
        log_action('create', 'assignment',
                   entity_name=f'{asset.asset_tag} → {emp.name}',
                   details=f'Activo: {asset.name} asignado a {emp.name} ({emp.employee_id})')
        db.session.commit()
        flash('Asignación creada correctamente.', 'success')
        return redirect(url_for('assignments_list'))
    return render_template('assignments/form.html',
                           assets=assets, employees=employees, pre_asset=pre_asset)


@app.route('/assignments/<int:id>/return', methods=['POST'])
@login_required
def assignment_return(id):
    assignment = Assignment.query.get_or_404(id)
    if assignment.returned_date:
        flash('Este activo ya fue devuelto.', 'warning')
        return redirect(url_for('assignments_list'))
    assignment.returned_date   = parse_date(request.form.get('returned_date')) or date.today()
    assignment.asset.status    = 'available'
    log_action('update', 'assignment',
               entity_id=assignment.id,
               entity_name=assignment.asset.name,
               details=f'Devolución de {assignment.asset.asset_tag} por {assignment.employee.name}')
    db.session.commit()
    flash(f'Activo "{assignment.asset.name}" devuelto correctamente.', 'success')
    return redirect(url_for('assignments_list'))


# ── Maintenance ───────────────────────────────────────────────────────────────

_MAINT_DIR = os.path.join(os.path.dirname(__file__), 'instance', 'uploads', 'maintenance')
os.makedirs(_MAINT_DIR, exist_ok=True)


@app.route('/maintenance')
@login_required
def maintenance_list():
    status_filter = request.args.get('status', '')
    type_filter   = request.args.get('type', '')
    q = Maintenance.query
    if status_filter:
        q = q.filter_by(status=status_filter)
    if type_filter:
        q = q.filter_by(maintenance_type=type_filter)
    records = q.order_by(Maintenance.created_at.desc()).all()
    counts = {
        'total':      Maintenance.query.count(),
        'pendiente':  Maintenance.query.filter_by(status='pendiente').count(),
        'en_proceso': Maintenance.query.filter_by(status='en_proceso').count(),
        'completado': Maintenance.query.filter_by(status='completado').count(),
        'cerrado':    Maintenance.query.filter_by(status='cerrado').count(),
    }
    return render_template('maintenance/list.html',
                           records=records, counts=counts,
                           status_filter=status_filter, type_filter=type_filter)


@app.route('/maintenance/new', methods=['GET', 'POST'])
@login_required
def maintenance_new():
    assets = Asset.query.filter(Asset.status != 'disposed').order_by(Asset.name).all()
    pre_asset_id = request.args.get('asset_id', type=int)

    if request.method == 'POST':
        asset_id = request.form.get('asset_id', type=int)
        asset = Asset.query.get_or_404(asset_id)

        import json as _jm
        from datetime import datetime as _dt

        # Build action plan from dynamic rows
        tasks        = request.form.getlist('task[]')
        responsibles = request.form.getlist('task_responsible[]')
        deadlines    = request.form.getlist('task_deadline[]')
        plan = [{'task': t, 'responsible': r, 'deadline': d}
                for t, r, d in zip(tasks, responsibles, deadlines) if t.strip()]

        m = Maintenance(
            asset_id            = asset_id,
            ticket_folio        = request.form.get('ticket_folio', '').strip() or f'MNT-{date.today().strftime("%Y%m%d")}-{asset_id}',
            maintenance_type    = request.form.get('maintenance_type', 'correctivo'),
            status              = 'en_proceso',
            prev_asset_status   = asset.status,
            reported_date       = parse_date(request.form.get('reported_date')) or date.today(),
            reported_by         = request.form.get('reported_by', '').strip(),
            process_name        = request.form.get('process_name', '').strip(),
            process_responsible = request.form.get('process_responsible', '').strip(),
            nc_source           = request.form.get('nc_source', ''),
            description         = request.form.get('description', '').strip(),
            analysis_method     = request.form.get('analysis_method', '').strip(),
            participants        = request.form.get('participants', '').strip(),
            root_cause_analysis = request.form.get('root_cause_analysis', '').strip(),
            root_cause          = request.form.get('root_cause', '').strip(),
            correction_desc     = request.form.get('correction_desc', '').strip(),
            action_plan         = _jm.dumps(plan, ensure_ascii=False),
            proposed_close_date = parse_date(request.form.get('proposed_close_date')),
            followup_responsible= request.form.get('followup_responsible', '').strip(),
            close_responsible   = request.form.get('close_responsible', '').strip(),
            notes               = request.form.get('notes', '').strip(),
        )
        # Change asset status to maintenance
        asset.status = 'maintenance'
        asset.last_maintenance = date.today()

        db.session.add(m)
        db.session.flush()  # get m.id

        # Handle photo uploads
        _maintenance_save_photos(m, request.files.getlist('photos_antes'), 'antes')
        _maintenance_save_photos(m, request.files.getlist('photos_despues'), 'despues')

        log_action('create', 'maintenance', entity_id=m.id,
                   entity_name=f'{m.ticket_folio} — {asset.name}',
                   details=f'Tipo: {m.maintenance_type} | Activo: {asset.asset_tag}')
        db.session.commit()
        flash(f'Ticket {m.ticket_folio} creado. Activo "{asset.name}" marcado como En Mantenimiento.', 'success')
        return redirect(url_for('maintenance_detail', id=m.id))

    return render_template('maintenance/form.html',
                           record=None, assets=assets, pre_asset_id=pre_asset_id,
                           type_choices=Maintenance.TYPE_CHOICES,
                           nc_sources=Maintenance.NC_SOURCES)


def _maintenance_save_photos(record, files, photo_type):
    """Save uploaded photos to disk and append to record.photos JSON."""
    import json as _jm, uuid
    current = record.photos_list
    for f in files:
        if not f or not f.filename:
            continue
        ext  = os.path.splitext(f.filename)[1].lower()
        if ext not in ('.jpg', '.jpeg', '.png', '.webp', '.gif', '.heic'):
            continue
        fname = f'{record.id}_{photo_type}_{uuid.uuid4().hex[:8]}{ext}'
        fpath = os.path.join(_MAINT_DIR, fname)
        f.save(fpath)
        current.append({'path': fname, 'name': f.filename,
                        'photo_type': photo_type, 'caption': ''})
    record.photos = _jm.dumps(current, ensure_ascii=False)


@app.route('/maintenance/<int:id>')
@login_required
def maintenance_detail(id):
    m = Maintenance.query.get_or_404(id)
    return render_template('maintenance/detail.html', record=m,
                           type_choices=Maintenance.TYPE_CHOICES,
                           nc_sources=Maintenance.NC_SOURCES)


@app.route('/maintenance/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def maintenance_edit(id):
    m    = Maintenance.query.get_or_404(id)
    assets = Asset.query.filter(Asset.status != 'disposed').order_by(Asset.name).all()

    if request.method == 'POST':
        import json as _jm
        tasks        = request.form.getlist('task[]')
        responsibles = request.form.getlist('task_responsible[]')
        deadlines    = request.form.getlist('task_deadline[]')
        plan = [{'task': t, 'responsible': r, 'deadline': d}
                for t, r, d in zip(tasks, responsibles, deadlines) if t.strip()]

        eff = request.form.get('effectiveness_ok')

        m.ticket_folio         = request.form.get('ticket_folio', m.ticket_folio).strip()
        m.maintenance_type     = request.form.get('maintenance_type', m.maintenance_type)
        m.reported_date        = parse_date(request.form.get('reported_date')) or m.reported_date
        m.reported_by          = request.form.get('reported_by', '').strip()
        m.process_name         = request.form.get('process_name', '').strip()
        m.process_responsible  = request.form.get('process_responsible', '').strip()
        m.nc_source            = request.form.get('nc_source', '')
        m.description          = request.form.get('description', '').strip()
        m.analysis_method      = request.form.get('analysis_method', '').strip()
        m.participants         = request.form.get('participants', '').strip()
        m.root_cause_analysis  = request.form.get('root_cause_analysis', '').strip()
        m.root_cause           = request.form.get('root_cause', '').strip()
        m.correction_desc      = request.form.get('correction_desc', '').strip()
        m.action_plan          = _jm.dumps(plan, ensure_ascii=False)
        m.proposed_close_date  = parse_date(request.form.get('proposed_close_date'))
        m.followup_responsible = request.form.get('followup_responsible', '').strip()
        m.close_responsible    = request.form.get('close_responsible', '').strip()
        m.effectiveness_ok     = True if eff == '1' else (False if eff == '0' else None)
        m.effectiveness_notes  = request.form.get('effectiveness_notes', '').strip()
        m.notes                = request.form.get('notes', '').strip()
        m.updated_at           = datetime.utcnow()

        # New photos
        _maintenance_save_photos(m, request.files.getlist('photos_antes'), 'antes')
        _maintenance_save_photos(m, request.files.getlist('photos_despues'), 'despues')

        log_action('update', 'maintenance', entity_id=m.id,
                   entity_name=m.ticket_folio,
                   details=f'Mantenimiento actualizado')
        db.session.commit()
        flash('Ticket de mantenimiento actualizado.', 'success')
        return redirect(url_for('maintenance_detail', id=m.id))

    return render_template('maintenance/form.html',
                           record=m, assets=assets, pre_asset_id=m.asset_id,
                           type_choices=Maintenance.TYPE_CHOICES,
                           nc_sources=Maintenance.NC_SOURCES)


@app.route('/maintenance/<int:id>/status', methods=['POST'])
@login_required
def maintenance_status(id):
    m      = Maintenance.query.get_or_404(id)
    new_st = request.form.get('status')
    if new_st not in [s for s, _ in Maintenance.STATUS_CHOICES]:
        flash('Estado inválido.', 'danger')
        return redirect(url_for('maintenance_detail', id=id))

    m.status = new_st
    if new_st == 'cerrado':
        m.actual_close_date = date.today()
        # Restore asset to previous status (or 'available' if unknown)
        restore = m.prev_asset_status or 'available'
        if restore == 'maintenance':
            restore = 'available'
        m.asset.status = restore
        flash(f'Ticket cerrado. Activo "{m.asset.name}" restaurado a "{restore}".', 'success')
    elif new_st == 'en_proceso':
        m.asset.status = 'maintenance'
        flash('Ticket marcado como En Proceso.', 'info')
    else:
        flash(f'Estado actualizado a {m.status_label}.', 'info')

    log_action('update', 'maintenance', entity_id=m.id,
               entity_name=m.ticket_folio,
               details=f'Estado → {new_st}')
    db.session.commit()
    return redirect(url_for('maintenance_detail', id=id))


@app.route('/maintenance/<int:id>/upload-document', methods=['POST'])
@login_required
def maintenance_upload_doc(id):
    m = Maintenance.query.get_or_404(id)
    f = request.files.get('document')
    if not f or not f.filename:
        flash('No se seleccionó ningún archivo.', 'warning')
        return redirect(url_for('maintenance_detail', id=id))
    import uuid
    ext   = os.path.splitext(f.filename)[1].lower()
    fname = f'doc_{id}_{uuid.uuid4().hex[:8]}{ext}'
    fpath = os.path.join(_MAINT_DIR, fname)
    f.save(fpath)
    m.document_path = fname
    m.document_name = f.filename
    db.session.commit()
    flash('Formato de mantenimiento subido correctamente.', 'success')
    return redirect(url_for('maintenance_detail', id=id))


@app.route('/maintenance/<int:id>/document')
@login_required
def maintenance_download_doc(id):
    m = Maintenance.query.get_or_404(id)
    if not m.document_path:
        flash('No hay documento adjunto.', 'warning')
        return redirect(url_for('maintenance_detail', id=id))
    fpath = os.path.join(_MAINT_DIR, m.document_path)
    return send_file(fpath, download_name=m.document_name or m.document_path, as_attachment=True)


@app.route('/maintenance/<int:id>/photo/<int:idx>')
@login_required
def maintenance_photo(id, idx):
    m = Maintenance.query.get_or_404(id)
    photos = m.photos_list
    if idx >= len(photos):
        return 'Not found', 404
    fpath = os.path.join(_MAINT_DIR, photos[idx]['path'])
    return send_file(fpath)


@app.route('/maintenance/<int:id>/delete-photo/<int:idx>', methods=['POST'])
@login_required
def maintenance_delete_photo(id, idx):
    import json as _jm
    m      = Maintenance.query.get_or_404(id)
    photos = m.photos_list
    if 0 <= idx < len(photos):
        removed = photos.pop(idx)
        try:
            os.remove(os.path.join(_MAINT_DIR, removed['path']))
        except OSError:
            pass
        m.photos = _jm.dumps(photos, ensure_ascii=False)
        db.session.commit()
    return redirect(url_for('maintenance_detail', id=id))


@app.route('/maintenance/<int:id>/pdf')
@login_required
def maintenance_pdf(id):
    """Genera el FO-SGSI-20 pre-llenado con los datos del ticket."""
    from maintenance_pdf import generate_fo_sgsi20
    m   = Maintenance.query.get_or_404(id)
    buf = generate_fo_sgsi20(m)
    filename = f'FO-SGSI-20_{m.ticket_folio}_{date.today().strftime("%Y%m%d")}.pdf'
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/pdf')


@app.route('/maintenance/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def maintenance_delete(id):
    m = Maintenance.query.get_or_404(id)
    name = m.ticket_folio
    # Restore asset status
    if m.asset.status == 'maintenance':
        m.asset.status = m.prev_asset_status or 'available'
    db.session.delete(m)
    log_action('delete', 'maintenance', entity_id=id, entity_name=name)
    db.session.commit()
    flash(f'Ticket "{name}" eliminado.', 'warning')
    return redirect(url_for('maintenance_list'))


# ── Licenses ──────────────────────────────────────────────────────────────────

_LIC_UPLOAD_DIR = os.path.join(os.path.dirname(__file__), 'instance', 'uploads', 'licenses')
os.makedirs(_LIC_UPLOAD_DIR, exist_ok=True)


@app.route('/licenses')
@login_required
def license_list():
    q        = request.args.get('q', '').strip()
    fstatus  = request.args.get('status', '')
    fvendor  = request.args.get('vendor', '')
    fcateg   = request.args.get('category', '')

    query = License.query
    if q:
        like = f'%{q}%'
        query = query.filter(
            db.or_(License.name.ilike(like), License.vendor.ilike(like),
                   License.software.ilike(like), License.tenant_name.ilike(like))
        )
    if fvendor:
        query = query.filter(License.vendor.ilike(f'%{fvendor}%'))
    if fcateg:
        query = query.filter(License.category == fcateg)

    all_lics = query.order_by(License.vendor, License.name).all()

    # Filter by effective_status after load (computed property)
    if fstatus:
        all_lics = [l for l in all_lics if l.effective_status == fstatus]

    # Stats
    today = date.today()
    soon  = today + timedelta(days=30)
    total      = License.query.count()
    active_c   = sum(1 for l in License.query.all() if l.effective_status == 'active')
    expiring_c = sum(1 for l in License.query.all() if l.effective_status == 'expiring')
    expired_c  = sum(1 for l in License.query.all() if l.effective_status == 'expired')

    vendors    = db.session.query(License.vendor).filter(License.vendor != None).distinct().order_by(License.vendor).all()
    vendors    = [v[0] for v in vendors if v[0]]

    return render_template('licenses/list.html',
                           licenses=all_lics,
                           q=q, fstatus=fstatus, fvendor=fvendor, fcateg=fcateg,
                           total=total, active_c=active_c,
                           expiring_c=expiring_c, expired_c=expired_c,
                           vendors=vendors,
                           category_choices=License.CATEGORY_CHOICES)


@app.route('/licenses/new', methods=['GET', 'POST'])
@login_required
def license_new():
    if request.method == 'POST':
        lic = License(
            name          = request.form.get('name', '').strip(),
            vendor        = request.form.get('vendor', '').strip() or None,
            software      = request.form.get('software', '').strip() or None,
            category      = request.form.get('category') or None,
            license_type  = request.form.get('license_type', 'subscription'),
            license_key   = request.form.get('license_key', '').strip() or None,
            is_microsoft  = bool(request.form.get('is_microsoft')),
            tenant_id     = request.form.get('tenant_id', '').strip() or None,
            tenant_name   = request.form.get('tenant_name', '').strip() or None,
            tenant_domain = request.form.get('tenant_domain', '').strip() or None,
            subscription_id = request.form.get('subscription_id', '').strip() or None,
            sku_name      = request.form.get('sku_name', '').strip() or None,
            seat_count    = int(request.form['seat_count']) if request.form.get('seat_count') else None,
            purchase_cost = float(request.form['purchase_cost']) if request.form.get('purchase_cost') else None,
            renewal_cost  = float(request.form['renewal_cost']) if request.form.get('renewal_cost') else None,
            currency      = request.form.get('currency', 'MXN'),
            purchase_date = _parse_date(request.form.get('purchase_date')),
            expiry_date   = _parse_date(request.form.get('expiry_date')),
            renewal_date  = _parse_date(request.form.get('renewal_date')),
            status        = request.form.get('status', 'active'),
            notes         = request.form.get('notes', '').strip() or None,
        )
        if not lic.name:
            flash('El nombre es requerido.', 'danger')
            return render_template('licenses/form.html', record=None,
                                   category_choices=License.CATEGORY_CHOICES,
                                   type_choices=License.TYPE_CHOICES,
                                   currency_choices=License.CURRENCY_CHOICES,
                                   form=request.form)
        db.session.add(lic)
        log_action('create', 'license', entity_name=lic.name,
                   details=f'Licencia creada: {lic.vendor} {lic.name}')
        db.session.commit()
        flash(f'Licencia "{lic.name}" creada.', 'success')
        return redirect(url_for('license_detail', id=lic.id))

    return render_template('licenses/form.html', record=None,
                           category_choices=License.CATEGORY_CHOICES,
                           type_choices=License.TYPE_CHOICES,
                           currency_choices=License.CURRENCY_CHOICES,
                           form={})


@app.route('/licenses/<int:id>')
@login_required
def license_detail(id):
    lic       = License.query.get_or_404(id)
    employees = Employee.query.filter_by(active=True).order_by(Employee.name).all()
    assets    = Asset.query.filter(Asset.status.notin_(['retired', 'disposed'])).order_by(Asset.name).all()
    return render_template('licenses/detail.html', record=lic,
                           employees=employees, assets=assets,
                           category_choices=License.CATEGORY_CHOICES,
                           type_choices=License.TYPE_CHOICES)


@app.route('/licenses/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def license_edit(id):
    lic = License.query.get_or_404(id)
    if request.method == 'POST':
        lic.name          = request.form.get('name', '').strip()
        lic.vendor        = request.form.get('vendor', '').strip() or None
        lic.software      = request.form.get('software', '').strip() or None
        lic.category      = request.form.get('category') or None
        lic.license_type  = request.form.get('license_type', 'subscription')
        lic.license_key   = request.form.get('license_key', '').strip() or None
        lic.is_microsoft  = bool(request.form.get('is_microsoft'))
        lic.tenant_id     = request.form.get('tenant_id', '').strip() or None
        lic.tenant_name   = request.form.get('tenant_name', '').strip() or None
        lic.tenant_domain = request.form.get('tenant_domain', '').strip() or None
        lic.subscription_id = request.form.get('subscription_id', '').strip() or None
        lic.sku_name      = request.form.get('sku_name', '').strip() or None
        lic.seat_count    = int(request.form['seat_count']) if request.form.get('seat_count') else None
        lic.purchase_cost = float(request.form['purchase_cost']) if request.form.get('purchase_cost') else None
        lic.renewal_cost  = float(request.form['renewal_cost']) if request.form.get('renewal_cost') else None
        lic.currency      = request.form.get('currency', 'MXN')
        lic.purchase_date = _parse_date(request.form.get('purchase_date'))
        lic.expiry_date   = _parse_date(request.form.get('expiry_date'))
        lic.renewal_date  = _parse_date(request.form.get('renewal_date'))
        lic.status        = request.form.get('status', 'active')
        lic.notes         = request.form.get('notes', '').strip() or None
        lic.updated_at    = datetime.utcnow()
        log_action('update', 'license', entity_id=lic.id, entity_name=lic.name)
        db.session.commit()
        flash('Licencia actualizada.', 'success')
        return redirect(url_for('license_detail', id=lic.id))

    return render_template('licenses/form.html', record=lic,
                           category_choices=License.CATEGORY_CHOICES,
                           type_choices=License.TYPE_CHOICES,
                           currency_choices=License.CURRENCY_CHOICES,
                           form={})


@app.route('/licenses/<int:id>/assign', methods=['POST'])
@login_required
def license_assign(id):
    lic  = License.query.get_or_404(id)
    # Check seat availability
    if lic.seat_count is not None and lic.used_seats >= lic.seat_count:
        flash('No hay asientos disponibles en esta licencia.', 'danger')
        return redirect(url_for('license_detail', id=id))

    emp_id   = request.form.get('employee_id') or None
    asset_id = request.form.get('asset_id')    or None
    if not emp_id and not asset_id:
        flash('Selecciona un empleado o un activo para asignar.', 'warning')
        return redirect(url_for('license_detail', id=id))

    asn = LicenseAssignment(
        license_id    = lic.id,
        employee_id   = int(emp_id)   if emp_id   else None,
        asset_id      = int(asset_id) if asset_id else None,
        assigned_date = _parse_date(request.form.get('assigned_date')) or date.today(),
        notes         = request.form.get('notes', '').strip() or None,
    )
    db.session.add(asn)
    target = Employee.query.get(emp_id).name if emp_id else Asset.query.get(asset_id).name
    log_action('create', 'license_assignment', entity_id=lic.id, entity_name=lic.name,
               details=f'Asignada a: {target}')
    db.session.commit()
    flash(f'Licencia asignada a {target}.', 'success')
    return redirect(url_for('license_detail', id=id))


@app.route('/licenses/<int:id>/unassign/<int:aid>', methods=['POST'])
@login_required
def license_unassign(id, aid):
    asn = LicenseAssignment.query.get_or_404(aid)
    lic = License.query.get_or_404(id)
    target = asn.employee.name if asn.employee else (asn.asset.name if asn.asset else '?')
    db.session.delete(asn)
    log_action('delete', 'license_assignment', entity_id=lic.id, entity_name=lic.name,
               details=f'Remoción de asignación: {target}')
    db.session.commit()
    flash(f'Asignación de {target} removida.', 'info')
    return redirect(url_for('license_detail', id=id))


@app.route('/licenses/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def license_delete(id):
    lic  = License.query.get_or_404(id)
    name = lic.name
    db.session.delete(lic)
    log_action('delete', 'license', entity_id=id, entity_name=name)
    db.session.commit()
    flash(f'Licencia "{name}" eliminada.', 'warning')
    return redirect(url_for('license_list'))


# ── Categories ────────────────────────────────────────────────────────────────

@app.route('/categories')
@login_required
def categories_list():
    categories = Category.query.order_by(Category.name).all()
    return render_template('categories/list.html', categories=categories)


@app.route('/categories/new', methods=['GET', 'POST'])
@login_required
def category_new():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if Category.query.filter_by(name=name).first():
            flash(f'La categoría "{name}" ya existe.', 'danger')
            return render_template('categories/form.html', category=None, form=request.form)
        cat = Category(name=name,
                       description=request.form.get('description', '').strip() or None)
        db.session.add(cat)
        log_action('create', 'category', entity_name=name,
                   details=f'Categoría creada: {name}')
        db.session.commit()
        flash(f'Categoría "{cat.name}" creada.', 'success')
        return redirect(url_for('categories_list'))
    return render_template('categories/form.html', category=None, form={})


@app.route('/categories/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def category_edit(id):
    cat = Category.query.get_or_404(id)
    if request.method == 'POST':
        name     = request.form.get('name', '').strip()
        existing = Category.query.filter_by(name=name).first()
        if existing and existing.id != cat.id:
            flash(f'La categoría "{name}" ya existe.', 'danger')
            return render_template('categories/form.html', category=cat, form=request.form)
        old_name = cat.name
        cat.name        = name
        cat.description = request.form.get('description', '').strip() or None
        log_action('update', 'category', entity_id=cat.id, entity_name=name,
                   details=f'Categoría renombrada: {old_name} → {name}' if old_name != name else 'Descripción actualizada')
        db.session.commit()
        flash(f'Categoría "{cat.name}" actualizada.', 'success')
        return redirect(url_for('categories_list'))
    return render_template('categories/form.html', category=cat, form={})


@app.route('/categories/<int:id>/delete', methods=['POST'])
@login_required
def category_delete(id):
    cat = Category.query.get_or_404(id)
    if cat.assets:
        flash(f'No se puede eliminar "{cat.name}": tiene activos asociados.', 'danger')
        return redirect(url_for('categories_list'))
    name = cat.name
    log_action('delete', 'category', entity_id=cat.id, entity_name=name,
               details=f'Categoría eliminada: {name}')
    db.session.delete(cat)
    db.session.commit()
    flash(f'Categoría "{name}" eliminada.', 'warning')
    return redirect(url_for('categories_list'))


# ── Shipments (DHL / Foráneo) ─────────────────────────────────────────────────

@app.route('/shipments')
@login_required
def shipments_list():
    status_filter = request.args.get('status', '')
    query = Shipment.query
    if status_filter:
        query = query.filter_by(status=status_filter)
    page       = request.args.get('page', 1, type=int)
    pagination = query.order_by(Shipment.created_at.desc()).paginate(page=page, per_page=50, error_out=False)
    shipments  = pagination.items
    return render_template('shipments/list.html',
                           shipments=shipments, status_filter=status_filter,
                           pagination=pagination)


@app.route('/shipments/new', methods=['GET', 'POST'])
@app.route('/shipments/new/<int:asset_id>', methods=['GET', 'POST'])
@login_required
def shipment_new(asset_id=None):
    foraneo_assets = Asset.query.filter(
        Asset.location_type.in_(['foraneo', 'hibrido'])
    ).order_by(Asset.name).all()
    pre_asset = asset_id or request.args.get('asset_id')
    if request.method == 'POST':
        shipment = Shipment(
            asset_id=int(request.form.get('asset_id')),
            direction=request.form.get('direction', 'outbound'),
            carrier=request.form.get('carrier', 'DHL').strip(),
            tracking_number=request.form.get('tracking_number', '').strip(),
            origin=request.form.get('origin', '').strip() or None,
            destination=request.form.get('destination', '').strip() or None,
            recipient_name=request.form.get('recipient_name', '').strip() or None,
            status=request.form.get('status', 'pendiente'),
            ship_date=parse_date(request.form.get('ship_date')),
            estimated_delivery=parse_date(request.form.get('estimated_delivery')),
            notes=request.form.get('notes', '').strip() or None,
        )
        db.session.add(shipment)
        log_action('create', 'shipment',
                   entity_name=f'{shipment.carrier} #{shipment.tracking_number}',
                   details=f'Carrier: {shipment.carrier} | Tracking: {shipment.tracking_number} | Destino: {shipment.destination}')
        db.session.commit()
        # Auto-register in AfterShip
        import tracking as trk
        if trk._API_KEY and shipment.tracking_number:
            try:
                trk.refresh_shipment(shipment)
                db.session.commit()
            except Exception:
                pass
        flash(f'Envío {shipment.carrier} #{shipment.tracking_number} registrado.', 'success')
        return redirect(url_for('shipment_detail', id=shipment.id))
    return render_template('shipments/form.html',
                           foraneo_assets=foraneo_assets, shipment=None,
                           pre_asset=str(pre_asset) if pre_asset else '',
                           pre_direction='outbound', pre_origin='', form={})


@app.route('/shipments/<int:id>')
@login_required
def shipment_detail(id):
    shipment = Shipment.query.get_or_404(id)
    return render_template('shipments/detail.html', shipment=shipment)


@app.route('/shipments/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def shipment_edit(id):
    shipment       = Shipment.query.get_or_404(id)
    foraneo_assets = Asset.query.filter(
        Asset.location_type.in_(['foraneo', 'hibrido'])
    ).order_by(Asset.name).all()
    if request.method == 'POST':
        old_status = shipment.status
        shipment.carrier           = request.form.get('carrier', 'DHL').strip()
        shipment.tracking_number   = request.form.get('tracking_number', '').strip()
        shipment.origin            = request.form.get('origin', '').strip() or None
        shipment.destination       = request.form.get('destination', '').strip() or None
        shipment.recipient_name    = request.form.get('recipient_name', '').strip() or None
        shipment.status            = request.form.get('status', 'pendiente')
        shipment.ship_date         = parse_date(request.form.get('ship_date'))
        shipment.estimated_delivery = parse_date(request.form.get('estimated_delivery'))
        shipment.actual_delivery   = parse_date(request.form.get('actual_delivery'))
        shipment.notes             = request.form.get('notes', '').strip() or None
        log_action('update', 'shipment', entity_id=shipment.id,
                   entity_name=f'{shipment.carrier} #{shipment.tracking_number}',
                   details=f'Estado: {old_status} → {shipment.status}' if old_status != shipment.status else 'Actualización de datos')
        db.session.commit()
        flash(f'Envío #{shipment.tracking_number} actualizado.', 'success')
        return redirect(url_for('shipment_detail', id=shipment.id))
    return render_template('shipments/form.html',
                           foraneo_assets=foraneo_assets, shipment=shipment,
                           pre_asset=str(shipment.asset_id), form={})


@app.route('/shipments/<int:id>/delete', methods=['POST'])
@login_required
def shipment_delete(id):
    shipment = Shipment.query.get_or_404(id)
    log_action('delete', 'shipment', entity_id=shipment.id,
               entity_name=f'{shipment.carrier} #{shipment.tracking_number}',
               details='Envío eliminado')
    db.session.delete(shipment)
    db.session.commit()
    flash('Envío eliminado.', 'warning')
    return redirect(url_for('shipments_list'))


# ── Package Tracking (AfterShip) ─────────────────────────────────────────────

@app.route('/shipments/<int:id>/track', methods=['POST'])
@login_required
@csrf.exempt   # llamado también vía fetch desde el detalle
def shipment_track(id):
    """Rastrear un envío ahora mismo y actualizar su estado."""
    import tracking as trk
    shipment = Shipment.query.get_or_404(id)

    if not trk._API_KEY:
        return jsonify({'ok': False, 'error': 'AFTERSHIP_API_KEY no configurado en .env'}), 503

    try:
        ok = trk.refresh_shipment(shipment)
    except trk.AfterShipRateLimitError as e:
        return jsonify({
            'ok': False,
            'error': 'Límite diario de AfterShip alcanzado (plan gratuito: 100 req/día). Intenta de nuevo mañana o actualiza el plan en aftership.com.',
            'code': 429,
        }), 429
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Error inesperado: {e}'}), 502

    if ok:
        db.session.commit()
        import json as _json
        events = []
        if shipment.tracking_events:
            try:
                events = _json.loads(shipment.tracking_events)
            except Exception:
                pass
        return jsonify({
            'ok':       True,
            'status':   shipment.status,
            'tag':      shipment.tracking_tag,
            'eta':      shipment.est_delivery_afship.isoformat() if shipment.est_delivery_afship else None,
            'updated':  shipment.last_tracking_at.strftime('%d/%m/%Y %H:%M') if shipment.last_tracking_at else None,
            'events':   events[-10:],
        })
    return jsonify({'ok': False, 'error': 'No se pudo obtener información del carrier'}), 502


@app.route('/shipments/return/<int:asset_id>', methods=['GET', 'POST'])
@login_required
def shipment_return(asset_id):
    """Create an inbound (return) shipment for an asset."""
    asset = Asset.query.get_or_404(asset_id)
    ca = asset.current_assignment  # to get employee info
    foraneo_assets = Asset.query.filter(Asset.location_type.in_(['foraneo', 'hibrido'])).order_by(Asset.name).all()

    if request.method == 'POST':
        shipment = Shipment(
            asset_id=asset_id,
            direction='inbound',
            carrier=request.form.get('carrier', 'DHL').strip(),
            tracking_number=request.form.get('tracking_number', '').strip(),
            origin=request.form.get('origin', '').strip() or None,
            destination=request.form.get('destination', '').strip() or None,
            recipient_name=request.form.get('recipient_name', '').strip() or None,
            status=request.form.get('status', 'pendiente'),
            ship_date=parse_date(request.form.get('ship_date')),
            estimated_delivery=parse_date(request.form.get('estimated_delivery')),
            notes=request.form.get('notes', '').strip() or None,
        )
        db.session.add(shipment)
        log_action('create', 'shipment', entity_name=f'RETURN {shipment.carrier} #{shipment.tracking_number}',
                   details=f'Devolución de {asset.asset_tag} | Carrier: {shipment.carrier}')
        db.session.commit()

        # Auto-register in AfterShip if API key is set
        import tracking as trk
        if trk._API_KEY and shipment.tracking_number:
            try:
                trk.refresh_shipment(shipment)
                db.session.commit()
            except Exception:
                pass

        flash(f'Devolución registrada: {shipment.carrier} #{shipment.tracking_number}', 'success')
        return redirect(url_for('shipment_detail', id=shipment.id))

    # Pre-fill: origin = employee location, destination = RTS office
    pre_origin = ''
    if ca and ca.employee:
        pre_origin = ca.employee.department or ''

    return render_template('shipments/form.html',
                           foraneo_assets=foraneo_assets,
                           shipment=None,
                           pre_asset=str(asset_id),
                           pre_direction='inbound',
                           pre_origin=pre_origin,
                           form={})


@app.route('/webhooks/aftership', methods=['POST'])
@csrf.exempt
def aftership_webhook():
    """AfterShip sends real-time status updates here."""
    import tracking as trk, json as _json
    try:
        payload = request.get_json(force=True) or {}
        tracking_data = payload.get('data', {}).get('tracking', payload.get('data', {}))

        if not tracking_data:
            return jsonify({'ok': True})

        tracking_number = tracking_data.get('tracking_number')
        if not tracking_number:
            return jsonify({'ok': True})

        # Find matching shipment
        shipment = Shipment.query.filter_by(tracking_number=tracking_number).first()
        if shipment:
            shipment.tracking_tag     = tracking_data.get('tag')
            shipment.last_tracking_at = datetime.utcnow()

            tag = tracking_data.get('tag', '')
            new_status = Shipment.AFTERSHIP_STATUS_MAP.get(tag)
            if new_status:
                shipment.status = new_status

            # Store events
            events = tracking_data.get('events', tracking_data.get('checkpoints', []))[-20:]
            normalized = [{'message': e.get('message', ''), 'location': e.get('location', ''), 'checkpoint_time': e.get('occurred_at', e.get('checkpoint_time', ''))} for e in events]
            shipment.tracking_events = _json.dumps(normalized, ensure_ascii=False)

            db.session.commit()
            app.logger.info("Webhook: shipment %s updated to %s", shipment.id, shipment.status)
    except Exception as e:
        app.logger.exception("Webhook error: %s", e)

    return jsonify({'ok': True})


@app.route('/shipments/refresh-all', methods=['POST'])
@login_required
def shipments_refresh_all():
    """Actualiza todos los envíos activos — llamado por el scheduler o manualmente."""
    import tracking as trk
    updated = trk.refresh_all_active(app)
    flash(f'Tracking actualizado: {updated} envío(s) sincronizados.', 'success')
    return redirect(url_for('shipments_list'))


# ── Clients (Empresas) ────────────────────────────────────────────────────────

@app.route('/clients')
@login_required
def clients_list():
    q            = request.args.get('q', '')
    loc_filter   = request.args.get('location_type', '')
    clients_q    = Client.query
    if q:
        clients_q = clients_q.filter(db.or_(
            Client.name.ilike(f'%{q}%'),
            Client.contact_name.ilike(f'%{q}%'),
            Client.email.ilike(f'%{q}%'),
            Client.country.ilike(f'%{q}%'),
            Client.city.ilike(f'%{q}%'),
        ))
    if loc_filter:
        clients_q = clients_q.filter_by(location_type=loc_filter)
    clients = clients_q.order_by(Client.name).all()
    return render_template('clients/list.html', clients=clients, q=q, loc_filter=loc_filter)


def _save_client_form(cli, form):
    """Read form fields into a Client object. Returns list of validation errors."""
    errors = []
    name = form.get('name', '').strip()
    if not name:
        errors.append('El nombre del cliente es obligatorio.')
        return errors
    duplicate = Client.query.filter_by(name=name).first()
    if duplicate and duplicate.id != getattr(cli, 'id', None):
        errors.append(f'El cliente "{name}" ya existe.')
        return errors
    cli.name          = name
    cli.location_type = form.get('location_type', 'local')
    cli.contact_name  = form.get('contact_name', '').strip() or None
    cli.email         = form.get('email', '').strip() or None
    cli.phone         = form.get('phone', '').strip() or None
    cli.country       = form.get('country', '').strip() or None
    cli.city          = form.get('city', '').strip() or None
    cli.address       = form.get('address', '').strip() or None
    cli.rfc           = form.get('rfc', '').strip() or None
    cli.industry      = form.get('industry', '').strip() or None
    cli.website       = form.get('website', '').strip() or None
    cli.notes         = form.get('notes', '').strip() or None
    cli.start_date    = parse_date(form.get('start_date'))
    if hasattr(cli, 'id') and cli.id:   # edit only
        cli.active    = 'active' in form
    return errors


@app.route('/clients/new', methods=['GET', 'POST'])
@login_required
def client_new():
    if request.method == 'POST':
        cli    = Client(active=True)
        errors = _save_client_form(cli, request.form)
        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('clients/form.html', client=None, form=request.form)
        db.session.add(cli)
        log_action('create', 'client', entity_name=cli.name,
                   details=f'Cliente {cli.location_type}: {cli.name}')
        db.session.commit()
        flash(f'Cliente "{cli.name}" creado correctamente.', 'success')
        return redirect(url_for('clients_list'))
    return render_template('clients/form.html', client=None, form={})


@app.route('/clients/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def client_edit(id):
    cli = Client.query.get_or_404(id)
    if request.method == 'POST':
        errors = _save_client_form(cli, request.form)
        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('clients/form.html', client=cli, form=request.form)
        log_action('update', 'client', entity_id=cli.id, entity_name=cli.name,
                   details='Datos actualizados')
        db.session.commit()
        flash(f'Cliente "{cli.name}" actualizado.', 'success')
        return redirect(url_for('clients_list'))
    return render_template('clients/form.html', client=cli, form={})


@app.route('/clients/<int:id>/delete', methods=['POST'])
@login_required
def client_delete(id):
    cli = Client.query.get_or_404(id)
    if cli.assets:
        flash(f'No se puede eliminar "{cli.name}": tiene activos asociados.', 'danger')
        return redirect(url_for('clients_list'))
    name = cli.name
    log_action('delete', 'client', entity_id=cli.id, entity_name=name)
    db.session.delete(cli)
    db.session.commit()
    flash(f'Cliente "{name}" eliminado.', 'warning')
    return redirect(url_for('clients_list'))


@app.route('/clients/import', methods=['GET', 'POST'])
@login_required
def client_import():
    """Import clients from an Excel file (.xlsx)."""
    if request.method == 'POST':
        f = request.files.get('excel_file')
        if not f or not f.filename.endswith(('.xlsx', '.xls')):
            flash('Selecciona un archivo Excel (.xlsx)', 'danger')
            return redirect(url_for('client_import'))
        try:
            import pandas as pd
            df = pd.read_excel(f, dtype=str)
            df.columns = [c.strip().lower().replace(' ', '_') for c in df.columns]
            required = {'nombre', 'name'}
            name_col = next((c for c in df.columns if c in required or 'nombre' in c or 'name' in c), None)
            if not name_col:
                flash('El archivo debe tener una columna "Nombre" o "Name".', 'danger')
                return redirect(url_for('client_import'))

            # Column aliases
            def _col(aliases):
                for a in aliases:
                    c = next((c for c in df.columns if a in c), None)
                    if c:
                        return c
                return None

            col_map = {
                'name':          name_col,
                'location_type': _col(['tipo', 'location', 'local', 'foraneo']),
                'contact_name':  _col(['contacto', 'contact']),
                'email':         _col(['email', 'correo']),
                'phone':         _col(['telefono', 'phone', 'tel']),
                'country':       _col(['pais', 'country']),
                'city':          _col(['ciudad', 'city']),
                'address':       _col(['direccion', 'address']),
                'rfc':           _col(['rfc', 'tax']),
                'industry':      _col(['giro', 'industry', 'sector']),
                'website':       _col(['web', 'website', 'url']),
                'start_date':    _col(['fecha', 'start', 'inicio']),
                'notes':         _col(['notas', 'notes', 'observ']),
            }

            created = updated = skipped = 0
            for _, row in df.iterrows():
                raw_name = str(row.get(col_map['name'], '') or '').strip()
                if not raw_name or raw_name.lower() == 'nan':
                    skipped += 1
                    continue

                def _v(field):
                    col = col_map.get(field)
                    if not col:
                        return None
                    val = str(row.get(col, '') or '').strip()
                    return None if val.lower() in ('', 'nan', 'none') else val

                # Normalise location_type
                loc_raw = (_v('location_type') or 'local').lower()
                loc = 'foraneo' if any(x in loc_raw for x in ['foraneo', 'foráneo', 'remote', 'remoto']) else 'local'

                # Start date
                start_d = None
                sd_str  = _v('start_date')
                if sd_str:
                    try:
                        import pandas as _pd
                        start_d = _pd.to_datetime(sd_str, dayfirst=True).date()
                    except Exception:
                        pass

                existing = Client.query.filter_by(name=raw_name).first()
                if existing:
                    existing.location_type = loc
                    if _v('contact_name'): existing.contact_name = _v('contact_name')
                    if _v('email'):        existing.email        = _v('email')
                    if _v('phone'):        existing.phone        = _v('phone')
                    if _v('country'):      existing.country      = _v('country')
                    if _v('city'):         existing.city         = _v('city')
                    if _v('address'):      existing.address      = _v('address')
                    if _v('rfc'):          existing.rfc          = _v('rfc')
                    if _v('industry'):     existing.industry     = _v('industry')
                    if _v('website'):      existing.website      = _v('website')
                    if _v('notes'):        existing.notes        = _v('notes')
                    if start_d:            existing.start_date   = start_d
                    updated += 1
                else:
                    cli = Client(
                        name=raw_name, active=True,
                        location_type=loc,
                        contact_name=_v('contact_name'),
                        email=_v('email'), phone=_v('phone'),
                        country=_v('country'), city=_v('city'),
                        address=_v('address'), rfc=_v('rfc'),
                        industry=_v('industry'), website=_v('website'),
                        notes=_v('notes'), start_date=start_d,
                    )
                    db.session.add(cli)
                    created += 1

            db.session.commit()
            log_action('import', 'client', details=f'Excel import: {created} creados, {updated} actualizados, {skipped} omitidos')
            flash(f'Importación completada: {created} nuevos, {updated} actualizados, {skipped} omitidos.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error al procesar el archivo: {e}', 'danger')
        return redirect(url_for('clients_list'))

    return render_template('clients/import.html')


@app.route('/clients/import/template')
@login_required
def client_import_template():
    """Download an Excel template for client import."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Clientes'
    headers = ['Nombre', 'Tipo', 'Contacto', 'Email', 'Telefono',
               'Pais', 'Ciudad', 'Direccion', 'RFC', 'Giro', 'Web', 'Fecha', 'Notas']
    example = ['Remote Team Solutions', 'foraneo', 'Juan García', 'juan@rts.com',
               '+52 55 1234 5678', 'México', 'CDMX', 'Insurgentes 123', 'ABC123456XYZ',
               'Tecnología', 'https://rts.com', '01/01/2024', 'Cliente prioritario']
    header_fill = PatternFill('solid', fgColor='1D6F42')
    header_font = Font(bold=True, color='FFFFFF')
    for col, (h, ex) in enumerate(zip(headers, example), start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[c.column_letter].width = max(len(h), len(str(ex))) + 4
        ws.cell(row=2, column=col, value=ex)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='plantilla_clientes.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── Purchase Orders & Invoices ────────────────────────────────────────────────

_DOCS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'instance', 'uploads', 'documents')
os.makedirs(_DOCS_DIR, exist_ok=True)

_ALLOWED_DOC_EXT = {'.pdf', '.xml', '.xlsx', '.xls', '.jpg', '.jpeg', '.png', '.zip'}

def _save_doc_file(file_storage):
    """Save an uploaded file to _DOCS_DIR. Returns (stored_path, original_name, mime)."""
    import uuid, mimetypes
    original = file_storage.filename
    ext = os.path.splitext(original)[1].lower()
    if ext not in _ALLOWED_DOC_EXT:
        raise ValueError(f'Tipo de archivo no permitido: {ext}')
    stored_name = f'{uuid.uuid4().hex}{ext}'
    full_path   = os.path.join(_DOCS_DIR, stored_name)
    file_storage.save(full_path)
    mime = mimetypes.guess_type(original)[0] or 'application/octet-stream'
    return full_path, original, mime


# ── API: list + upload POs ────────────────────────────────────────────────────

@app.route('/api/purchase-orders')
@login_required
def api_purchase_orders():
    pos = PurchaseOrder.query.order_by(PurchaseOrder.created_at.desc()).all()
    return jsonify([{
        'id': p.id, 'display': p.display,
        'number': p.number,
        'date': p.date.strftime('%d/%m/%Y') if p.date else '',
        'supplier': p.supplier_name or (p.supplier.name if p.supplier else ''),
        'amount': p.total_amount, 'currency': p.currency,
        'has_file': bool(p.file_path),
    } for p in pos])


@app.route('/api/purchase-orders/upload', methods=['POST'])
@login_required
def api_po_upload():
    number  = request.form.get('number', '').strip()
    if not number:
        return jsonify({'ok': False, 'error': 'El número de OC es obligatorio'}), 400
    po = PurchaseOrder(
        number=number,
        date=parse_date(request.form.get('date')),
        supplier_name=request.form.get('supplier_name', '').strip() or None,
        total_amount=float(request.form.get('total_amount')) if request.form.get('total_amount') else None,
        currency=request.form.get('currency', 'MXN'),
        notes=request.form.get('notes', '').strip() or None,
    )
    f = request.files.get('file')
    if f and f.filename:
        try:
            po.file_path, po.file_name, po.file_mime = _save_doc_file(f)
        except ValueError as e:
            return jsonify({'ok': False, 'error': str(e)}), 400
    db.session.add(po)
    db.session.commit()
    log_action('create', 'purchase_order', entity_id=po.id, entity_name=po.number)
    return jsonify({'ok': True, 'id': po.id, 'display': po.display})


@app.route('/api/invoices')
@login_required
def api_invoices():
    invs = Invoice.query.order_by(Invoice.created_at.desc()).all()
    return jsonify([{
        'id': i.id, 'display': i.display,
        'number': i.number,
        'date': i.date.strftime('%d/%m/%Y') if i.date else '',
        'supplier': i.supplier_name or (i.supplier.name if i.supplier else ''),
        'amount': i.total_amount, 'currency': i.currency,
        'has_file': bool(i.file_path),
    } for i in invs])


@app.route('/api/invoices/upload', methods=['POST'])
@login_required
def api_invoice_upload():
    number = request.form.get('number', '').strip()
    if not number:
        return jsonify({'ok': False, 'error': 'El número de factura es obligatorio'}), 400
    inv = Invoice(
        number=number,
        date=parse_date(request.form.get('date')),
        supplier_name=request.form.get('supplier_name', '').strip() or None,
        total_amount=float(request.form.get('total_amount')) if request.form.get('total_amount') else None,
        currency=request.form.get('currency', 'MXN'),
        notes=request.form.get('notes', '').strip() or None,
    )
    f = request.files.get('file')
    if f and f.filename:
        try:
            inv.file_path, inv.file_name, inv.file_mime = _save_doc_file(f)
        except ValueError as e:
            return jsonify({'ok': False, 'error': str(e)}), 400
    db.session.add(inv)
    db.session.commit()
    log_action('create', 'invoice', entity_id=inv.id, entity_name=inv.number)
    return jsonify({'ok': True, 'id': inv.id, 'display': inv.display})


@app.route('/documents/po/<int:id>')
@login_required
def document_po_view(id):
    po = PurchaseOrder.query.get_or_404(id)
    if not po.file_path or not os.path.exists(po.file_path):
        flash('Archivo no disponible.', 'warning')
        return redirect(request.referrer or url_for('assets_list'))
    return send_file(po.file_path, download_name=po.file_name,
                     mimetype=po.file_mime or 'application/octet-stream',
                     as_attachment=False)


@app.route('/documents/invoice/<int:id>')
@login_required
def document_invoice_view(id):
    inv = Invoice.query.get_or_404(id)
    if not inv.file_path or not os.path.exists(inv.file_path):
        flash('Archivo no disponible.', 'warning')
        return redirect(request.referrer or url_for('assets_list'))
    return send_file(inv.file_path, download_name=inv.file_name,
                     mimetype=inv.file_mime or 'application/octet-stream',
                     as_attachment=False)


# ── Template filters & context ────────────────────────────────────────────────

@app.template_filter('format_date')
def format_date(value):
    if not value:
        return '—'
    if isinstance(value, str):
        return value
    return value.strftime('%d/%m/%Y')


@app.template_filter('format_currency')
def format_currency(value):
    if value is None:
        return '—'
    return f'${value:,.2f} MXN'


# ── Import Assets from Excel ──────────────────────────────────────────────────

@app.route('/assets/import', methods=['GET', 'POST'])
@module_required('inventory')
def asset_import():
    if request.method == 'POST':
        f = request.files.get('file')
        if not f or not f.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload a valid Excel file (.xlsx).', 'danger')
            return redirect(url_for('asset_import'))

        import openpyxl
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            flash(f'Could not read file: {e}', 'danger')
            return redirect(url_for('asset_import'))

        if len(rows) < 2:
            flash('File appears empty (no data rows).', 'warning')
            return redirect(url_for('asset_import'))

        # Map header → column index (case-insensitive)
        header = [str(h).strip().lower() if h else '' for h in rows[0]]
        def col(name):
            try: return header.index(name.lower())
            except ValueError: return None

        cat_cache = {c.name.lower(): c for c in Category.query.all()}
        created = updated = skipped = 0
        errors  = []

        for i, row in enumerate(rows[1:], start=2):
            def v(name, default=''):
                idx = col(name)
                val = row[idx] if idx is not None and idx < len(row) else None
                return str(val).strip() if val is not None else default

            tag = v('asset tag')
            if not tag:
                errors.append(f'Row {i}: missing asset tag — skipped.')
                skipped += 1
                continue

            existing = Asset.query.filter_by(asset_tag=tag).first()
            cat_name = v('category')
            category = cat_cache.get(cat_name.lower()) if cat_name else None
            if cat_name and not category:
                category = Category(name=cat_name)
                db.session.add(category)
                db.session.flush()
                cat_cache[cat_name.lower()] = category

            fields = {
                'name':         v('name') or tag,
                'asset_tag':    tag,
                'serial_number': v('serial number') or None,
                'manufacturer': v('manufacturer') or None,
                'model':        v('model') or None,
                'cpu':          v('cpu') or None,
                'ram':          v('ram') or None,
                'os_version':   v('os version') or None,
                'location':     v('location') or None,
                'location_type': v('location type') or 'en_sitio',
                'supplier':     v('supplier') or None,
                'notes':        v('notes') or None,
                'status':       v('status') or 'available',
                'category_id':  category.id if category else None,
            }
            # Parse dates
            for df in ['purchase date', 'warranty expiry', 'last maintenance']:
                val = v(df)
                field_name = df.replace(' ', '_')
                if val:
                    try:
                        import datetime as _dt
                        if isinstance(val, _dt.date):
                            fields[field_name] = val
                        else:
                            fields[field_name] = _dt.datetime.strptime(val[:10], '%Y-%m-%d').date()
                    except Exception:
                        pass
            # Cost
            cost_val = v('purchase cost')
            if cost_val:
                try: fields['purchase_cost'] = float(str(cost_val).replace(',', '').replace('$', ''))
                except Exception: pass

            if existing:
                for k, val in fields.items():
                    if val is not None and val != '':
                        setattr(existing, k, val)
                log_action('update', 'asset', entity_id=existing.id,
                           entity_name=existing.asset_tag, details='Updated via Excel import')
                updated += 1
            else:
                asset = Asset(**{k: v_ for k, v_ in fields.items() if v_ is not None and v_ != ''})
                db.session.add(asset)
                log_action('create', 'asset', entity_name=tag, details='Created via Excel import')
                created += 1

        db.session.commit()
        msg = f'Import complete: {created} created, {updated} updated, {skipped} skipped.'
        if errors:
            msg += f' ({len(errors)} errors — check logs.)'
        flash(msg, 'success' if not errors else 'warning')
        return redirect(url_for('assets_list'))

    # Template shows download link for the import template
    categories = Category.query.order_by(Category.name).all()
    return render_template('assets/import.html', categories=categories)


@app.route('/assets/import/template.xlsx')
@module_required('inventory')
def asset_import_template():
    """Download a blank Excel template for importing assets."""
    import openpyxl, io
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Assets'
    from openpyxl.styles import PatternFill, Font, Alignment
    headers = [
        'Asset Tag', 'Name', 'Category', 'Status', 'Location Type', 'Location',
        'Manufacturer', 'Model', 'CPU', 'RAM', 'OS Version', 'Serial Number',
        'Purchase Date', 'Purchase Cost', 'Supplier', 'Warranty Expiry',
        'Last Maintenance', 'Notes',
    ]
    navy = PatternFill('solid', fgColor='233C6E')
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, value=h)
        c.fill = navy
        c.font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(len(h) + 4, 14)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'
    # Example row
    example = ['RTS-001', 'Sample Laptop', 'Laptops', 'available', 'en_sitio', 'Office',
               'Dell', 'Latitude 5540', 'Intel Core i7', '16 GB', 'Windows 11', 'SN123456',
               '2024-01-15', '25000', 'CompuMexicana', '2026-01-15', '2025-06-01', 'Example row']
    for ci, val in enumerate(example, 1):
        ws.cell(2, ci, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name='RTS_Asset_Import_Template.xlsx')


# ── Bulk Asset Operations ─────────────────────────────────────────────────────

@app.route('/assets/bulk', methods=['POST'])
@csrf.exempt
@module_required('inventory')
def asset_bulk():
    from flask import jsonify
    ids    = request.form.getlist('ids[]')
    action = request.form.get('action')
    if not ids or not action:
        return jsonify({'ok': False, 'error': 'Missing ids or action'})

    assets = Asset.query.filter(Asset.id.in_([int(i) for i in ids])).all()
    changed = 0

    if action in ('available', 'in_use', 'maintenance', 'retired', 'disposed'):
        for a in assets:
            a.status = action
            log_action('update', 'asset', entity_id=a.id, entity_name=a.asset_tag,
                       details=f'Bulk status → {action}')
            changed += 1
    elif action.startswith('dept:'):
        dept_id = int(action.split(':')[1]) or None
        for a in assets:
            a.department_id = dept_id
            log_action('update', 'asset', entity_id=a.id, entity_name=a.asset_tag,
                       details=f'Bulk dept → {dept_id}')
            changed += 1
    else:
        return jsonify({'ok': False, 'error': 'Unknown action'})

    db.session.commit()
    return jsonify({'ok': True, 'changed': changed})


@app.route('/search')
@csrf.exempt
@login_required
def global_search():
    from flask import jsonify
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify({'assets': [], 'employees': [], 'projects': []})

    nq = _nrm(q)
    results = {'assets': [], 'employees': [], 'projects': []}

    # Assets
    from models import User as _User
    sess_user = session.get('user', {})
    is_admin  = sess_user.get('role') == 'admin'
    db_user   = _User.query.get(sess_user.get('id')) if sess_user.get('id') else None
    dept_filter = None
    if not is_admin and db_user and db_user.department_id:
        dept_filter = db_user.department_id

    aq = Asset.query.filter(db.or_(
        _search_col(Asset.name, nq),        _search_col(Asset.asset_tag, nq),
        _search_col(Asset.serial_number, nq), _search_col(Asset.manufacturer, nq),
        _search_col(Asset.model, nq),
    ))
    if dept_filter:
        aq = aq.filter_by(department_id=dept_filter)
    for a in aq.limit(6).all():
        results['assets'].append({
            'title': a.asset_tag,
            'sub':   a.name + (' · ' + a.category.name if a.category else ''),
            'url':   url_for('asset_detail', id=a.id),
        })

    # Employees
    for e in Employee.query.filter(db.or_(
        _search_col(Employee.name, nq),        _search_col(Employee.employee_id, nq),
        _search_col(Employee.email, nq),       _search_col(Employee.department, nq),
    )).limit(5).all():
        results['employees'].append({
            'title': e.name,
            'sub':   e.employee_id + (f' · {e.department}' if e.department else ''),
            'url':   url_for('employee_edit', id=e.id),
        })

    # Projects (if module access)
    from models import Project
    if is_admin or 'projects' in sess_user.get('modules', []):
        for p in Project.query.filter(db.or_(
            _search_col(Project.name, nq), _search_col(Project.code, nq),
        )).limit(5).all():
            results['projects'].append({
                'title': p.name,
                'sub':   p.code + f' · {p.status.replace("_"," ").title()}',
                'url':   url_for('projects.project_detail', id=p.id),
            })

    return jsonify(results)


# ── User Profile ──────────────────────────────────────────────────────────────

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def user_profile():
    from models import User as _User
    uid  = session['user']['id']
    user = _User.query.get_or_404(uid)

    if request.method == 'POST':
        name  = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip() or None
        pwd   = request.form.get('password', '').strip()
        pwd2  = request.form.get('password2', '').strip()

        if not name:
            flash('Name cannot be empty.', 'danger')
            return render_template('profile.html', user=user)
        if pwd and pwd != pwd2:
            flash('Passwords do not match.', 'danger')
            return render_template('profile.html', user=user)

        user.name  = name
        user.email = email
        if pwd:
            user.set_password(pwd)
        # Refresh session
        session['user']['name']  = user.name
        session['user']['email'] = user.email
        session.modified = True

        log_action('update', 'user', entity_id=uid, entity_name=user.name,
                   details='Profile updated by user')
        db.session.commit()
        flash('Profile updated successfully.', 'success')
        return redirect(url_for('user_profile'))

    return render_template('profile.html', user=user)


@app.route('/lang/<code>')
@login_required
def set_lang(code):
    if code in SUPPORTED_LANGS:
        session['lang'] = code
    return redirect(request.referrer or url_for('portal'))


@app.context_processor
def inject_globals():
    lang       = session.get('lang', DEFAULT_LANG)
    other_lang = 'es' if lang == 'en' else 'en'
    u          = session.get('user')

    # Pending access-requests count (admins only — shown in topbar badge)
    pending_requests_count = 0
    if u and u.get('role') == 'admin':
        try:
            pending_requests_count = AccessRequest.query.filter_by(status='pending').count()
        except Exception:
            pass

    # Warranty alerts for topbar (assets expiring in ≤30 days)
    warranty_alerts = 0
    if u and (u.get('role') == 'admin' or 'inventory' in u.get('modules', [])):
        try:
            soon = date.today() + timedelta(days=30)
            warranty_alerts = Asset.query.filter(
                Asset.warranty_expiry != None,        # noqa: E711
                Asset.warranty_expiry <= soon,
                Asset.warranty_expiry >= date.today(),
                Asset.status.notin_(['retired', 'disposed'])
            ).count()
        except Exception:
            pass

    # License expiry alerts (expiring or expired, non-cancelled)
    license_alerts = 0
    if u and (u.get('role') == 'admin' or 'inventory' in u.get('modules', [])):
        try:
            soon = date.today() + timedelta(days=30)
            license_alerts = License.query.filter(
                License.status != 'cancelled',
                License.expiry_date != None,       # noqa: E711
                License.expiry_date <= soon,
            ).count()
        except Exception:
            pass

    return {
        'today': date.today(),
        'current_user': u,
        'logo_exists': _img_exists('logo.png'),
        'remoties_exists': _img_exists('remoties.png'),
        'favicon_exists': _img_exists('favicon.png'),
        'ALL_MODULES': ALL_MODULES,
        'T': get_translations(lang),
        'lang': lang,
        'other_lang': other_lang,
        'pending_requests_count': pending_requests_count,
        'warranty_alerts': warranty_alerts,
        'license_alerts': license_alerts,
        'STATUS_BADGES': {
            'available':   'success',
            'in_use':      'primary',
            'maintenance': 'warning',
            'retired':     'secondary',
            'disposed':    'dark',
        },
        'STATUS_LABELS': {
            'available':   'Disponible',
            'in_use':      'En Uso',
            'maintenance': 'Mantenimiento',
            'retired':     'Retirado',
            'disposed':    'Desechado',
        },
        'SHIPMENT_STATUS_LABELS': {
            'pendiente':    'Pendiente',
            'en_transito':  'En Tránsito',
            'en_aduana':    'En Aduana',
            'entregado':    'Entregado',
            'devuelto':     'Devuelto',
        },
        'SHIPMENT_STATUS_BADGES': {
            'pendiente':    'secondary',
            'en_transito':  'primary',
            'en_aduana':    'warning',
            'entregado':    'success',
            'devuelto':     'dark',
        },
    }


# ── SETUP / CATALOGS ──────────────────────────────────────────────────────────

import qrcode, io, base64

def _qr_b64(text):
    qr = qrcode.QRCode(version=1, box_size=4, border=2)
    qr.add_data(text)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    return base64.b64encode(buf.getvalue()).decode()


@app.route('/setup')
@admin_required
def setup_index():
    counts = {
        'suppliers':   Supplier.query.count(),
        'brands':      Brand.query.count(),
        'categories':  Category.query.count(),
        'clients':     Client.query.filter_by(active=True).count(),
        'departments': Department.query.filter_by(active=True).count(),
    }
    cfg = IDConfig.get()
    return render_template('setup/index.html', counts=counts, cfg=cfg)


# ── Suppliers ─────────────────────────────────────────────────────────────────

@app.route('/setup/suppliers')
@admin_required
def suppliers_list():
    suppliers = Supplier.query.order_by(Supplier.name).all()
    return render_template('setup/suppliers.html', suppliers=suppliers)


@app.route('/setup/suppliers/new', methods=['GET', 'POST'])
@admin_required
def supplier_new():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('El nombre del proveedor es obligatorio.', 'danger')
            return render_template('setup/supplier_form.html', supplier=None, form=request.form)
        if Supplier.query.filter_by(name=name).first():
            flash(f'Ya existe un proveedor con el nombre "{name}".', 'danger')
            return render_template('setup/supplier_form.html', supplier=None, form=request.form)
        s = Supplier(
            name=name,
            contact_name=request.form.get('contact_name', '').strip() or None,
            email=request.form.get('email', '').strip() or None,
            phone=request.form.get('phone', '').strip() or None,
            website=request.form.get('website', '').strip() or None,
            country=request.form.get('country', '').strip() or None,
            notes=request.form.get('notes', '').strip() or None,
            active='active' in request.form,
        )
        db.session.add(s)
        log_action('create', 'supplier', entity_name=name)
        db.session.commit()
        flash(f'Proveedor "{name}" creado.', 'success')
        return redirect(url_for('suppliers_list'))
    return render_template('setup/supplier_form.html', supplier=None, form={})


@app.route('/setup/suppliers/<int:id>/edit', methods=['GET', 'POST'])
@admin_required
def supplier_edit(id):
    s = Supplier.query.get_or_404(id)
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('El nombre del proveedor es obligatorio.', 'danger')
            return render_template('setup/supplier_form.html', supplier=s, form=request.form)
        dup = Supplier.query.filter_by(name=name).first()
        if dup and dup.id != s.id:
            flash(f'Ya existe un proveedor con el nombre "{name}".', 'danger')
            return render_template('setup/supplier_form.html', supplier=s, form=request.form)
        s.name         = name
        s.contact_name = request.form.get('contact_name', '').strip() or None
        s.email        = request.form.get('email', '').strip() or None
        s.phone        = request.form.get('phone', '').strip() or None
        s.website      = request.form.get('website', '').strip() or None
        s.country      = request.form.get('country', '').strip() or None
        s.notes        = request.form.get('notes', '').strip() or None
        s.active       = 'active' in request.form
        log_action('update', 'supplier', entity_id=s.id, entity_name=s.name)
        db.session.commit()
        flash(f'Proveedor "{s.name}" actualizado.', 'success')
        return redirect(url_for('suppliers_list'))
    return render_template('setup/supplier_form.html', supplier=s, form={})


@app.route('/setup/suppliers/<int:id>/delete', methods=['POST'])
@admin_required
def supplier_delete(id):
    s = Supplier.query.get_or_404(id)
    name = s.name
    log_action('delete', 'supplier', entity_id=s.id, entity_name=name)
    db.session.delete(s)
    db.session.commit()
    flash(f'Proveedor "{name}" eliminado.', 'success')
    return redirect(url_for('suppliers_list'))


# ── Absolute Integration ──────────────────────────────────────────────────────

def _get_absolute_client():
    """Devuelve un AbsoluteClient configurado o None si no hay credenciales."""
    from absolute import AbsoluteClient, AbsoluteAuthError
    tid    = AppSetting.get('absolute_token_id', '')
    tsecret= AppSetting.get('absolute_token_secret', '')
    if not tid or not tsecret:
        return None
    return AbsoluteClient(tid, tsecret)


@app.route('/setup/absolute', methods=['GET', 'POST'])
@admin_required
def setup_absolute():
    """Configuración de credenciales de Absolute."""
    from absolute import AbsoluteClient, AbsoluteError
    test_result = None

    if request.method == 'POST':
        action = request.form.get('action', 'save')
        tid    = request.form.get('token_id', '').strip()
        tsec   = request.form.get('token_secret', '').strip()

        if action == 'test':
            if not tid or not tsec:
                test_result = {'ok': False, 'message': 'Ingresa el Token ID y Token Secret antes de probar.'}
            else:
                try:
                    result = AbsoluteClient(tid, tsec).test_connection()
                    test_result = result
                except AbsoluteError as e:
                    test_result = {'ok': False, 'message': str(e)}
        else:
            AppSetting.set('absolute_token_id',     tid)
            AppSetting.set('absolute_token_secret', tsec)
            db.session.commit()
            flash('Credenciales de Absolute guardadas.', 'success')
            return redirect(url_for('setup_absolute'))

    configured = bool(AppSetting.get('absolute_token_id'))
    return render_template('setup/absolute.html',
                           token_id=AppSetting.get('absolute_token_id', ''),
                           token_secret=AppSetting.get('absolute_token_secret', ''),
                           configured=configured,
                           test_result=test_result)


@app.route('/assets/<int:id>/absolute/search-serial', methods=['POST'])
@login_required
def asset_absolute_search(id):
    """Busca el activo en Absolute por su número de serie y devuelve candidatos."""
    from absolute import AbsoluteError, parse_device
    asset = Asset.query.get_or_404(id)
    client = _get_absolute_client()
    if not client:
        return jsonify({'error': 'Absolute no está configurado. Ve a Setup → Absolute.'}), 400

    serial = asset.serial_number or ''
    name   = request.form.get('search_name', asset.name or '')

    results = []
    try:
        if serial:
            raw = client.search_by_serial(serial)
            results = [parse_device(d) for d in raw]
        if not results and name:
            raw = client.search_by_name(name)
            results = [parse_device(d) for d in raw]
    except AbsoluteError as e:
        return jsonify({'error': str(e)}), 400

    # Serializar datetimes para JSON
    for r in results:
        if r.get('last_seen'):
            r['last_seen'] = r['last_seen'].strftime('%d/%m/%Y %H:%M')
        r.pop('raw', None)

    return jsonify({'results': results})


@app.route('/assets/<int:id>/absolute/link', methods=['POST'])
@login_required
def asset_absolute_link(id):
    """Liga el activo a un Device UID de Absolute y hace el primer sync."""
    from absolute import AbsoluteError, parse_device
    asset = Asset.query.get_or_404(id)
    client = _get_absolute_client()
    if not client:
        flash('Absolute no está configurado. Ve a Setup → Absolute.', 'danger')
        return redirect(url_for('asset_detail', id=id))

    device_id = request.form.get('device_id', '').strip()
    if not device_id:
        flash('Selecciona un dispositivo antes de ligar.', 'warning')
        return redirect(url_for('asset_detail', id=id))

    try:
        raw = client.get_device(device_id)
        dev = parse_device(raw)
    except AbsoluteError as e:
        flash(f'Error al obtener el dispositivo: {e}', 'danger')
        return redirect(url_for('asset_detail', id=id))

    asset.absolute_id        = dev['id'] or device_id
    asset.absolute_status    = dev['status']
    asset.absolute_username  = dev['username']
    asset.absolute_last_seen = dev['last_seen']
    asset.absolute_sync_at   = datetime.utcnow()
    log_action('update', 'asset', entity_id=asset.id, entity_name=asset.name,
               details=f'Ligado a Absolute Device: {device_id}')
    db.session.commit()
    flash(f'Activo ligado a Absolute: {dev["name"] or device_id}', 'success')
    return redirect(url_for('asset_detail', id=id))


@app.route('/assets/<int:id>/absolute/sync', methods=['POST'])
@login_required
def asset_absolute_sync(id):
    """Sincroniza datos del activo desde Absolute."""
    from absolute import AbsoluteError, parse_device
    asset = Asset.query.get_or_404(id)
    if not asset.absolute_id:
        flash('Este activo no está ligado a Absolute.', 'warning')
        return redirect(url_for('asset_detail', id=id))

    client = _get_absolute_client()
    if not client:
        flash('Absolute no está configurado.', 'danger')
        return redirect(url_for('asset_detail', id=id))

    try:
        raw = client.get_device(asset.absolute_id)
        dev = parse_device(raw)
    except AbsoluteError as e:
        flash(f'Error al sincronizar con Absolute: {e}', 'danger')
        return redirect(url_for('asset_detail', id=id))

    asset.absolute_status    = dev['status']
    asset.absolute_username  = dev['username']
    asset.absolute_last_seen = dev['last_seen']
    asset.absolute_sync_at   = datetime.utcnow()
    db.session.commit()
    flash('Datos de Absolute actualizados.', 'success')
    return redirect(url_for('asset_detail', id=id))


@app.route('/assets/<int:id>/absolute/unlink', methods=['POST'])
@login_required
def asset_absolute_unlink(id):
    """Desliga el activo de Absolute (solo borra el vínculo local)."""
    asset = Asset.query.get_or_404(id)
    asset.absolute_id        = None
    asset.absolute_status    = None
    asset.absolute_username  = None
    asset.absolute_last_seen = None
    asset.absolute_sync_at   = None
    log_action('update', 'asset', entity_id=asset.id, entity_name=asset.name,
               details='Desvinculado de Absolute')
    db.session.commit()
    flash('Vínculo con Absolute eliminado.', 'info')
    return redirect(url_for('asset_detail', id=id))


# ── Brands ────────────────────────────────────────────────────────────────────

@app.route('/setup/brands')
@admin_required
def brands_list():
    brands = Brand.query.order_by(Brand.name).all()
    return render_template('setup/brands.html', brands=brands)


@app.route('/setup/brands/new', methods=['GET', 'POST'])
@admin_required
def brand_new():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('El nombre de la marca es obligatorio.', 'danger')
            return render_template('setup/brand_form.html', brand=None, form=request.form)
        if Brand.query.filter_by(name=name).first():
            flash(f'Ya existe una marca con el nombre "{name}".', 'danger')
            return render_template('setup/brand_form.html', brand=None, form=request.form)
        b = Brand(
            name=name,
            description=request.form.get('description', '').strip() or None,
            active='active' in request.form,
        )
        db.session.add(b)
        log_action('create', 'brand', entity_name=name)
        db.session.commit()
        flash(f'Marca "{name}" creada.', 'success')
        return redirect(url_for('brands_list'))
    return render_template('setup/brand_form.html', brand=None, form={})


@app.route('/setup/brands/<int:id>/edit', methods=['GET', 'POST'])
@admin_required
def brand_edit(id):
    b = Brand.query.get_or_404(id)
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('El nombre de la marca es obligatorio.', 'danger')
            return render_template('setup/brand_form.html', brand=b, form=request.form)
        dup = Brand.query.filter_by(name=name).first()
        if dup and dup.id != b.id:
            flash(f'Ya existe una marca con el nombre "{name}".', 'danger')
            return render_template('setup/brand_form.html', brand=b, form=request.form)
        b.name        = name
        b.description = request.form.get('description', '').strip() or None
        b.active      = 'active' in request.form
        log_action('update', 'brand', entity_id=b.id, entity_name=b.name)
        db.session.commit()
        flash(f'Marca "{b.name}" actualizada.', 'success')
        return redirect(url_for('brands_list'))
    return render_template('setup/brand_form.html', brand=b, form={})


@app.route('/setup/brands/<int:id>/delete', methods=['POST'])
@admin_required
def brand_delete(id):
    b = Brand.query.get_or_404(id)
    name = b.name
    log_action('delete', 'brand', entity_id=b.id, entity_name=name)
    db.session.delete(b)
    db.session.commit()
    flash(f'Marca "{name}" eliminada.', 'success')
    return redirect(url_for('brands_list'))


# ── ID Config ─────────────────────────────────────────────────────────────────

@app.route('/setup/id-config', methods=['GET', 'POST'])
@admin_required
def id_config_view():
    cfg = IDConfig.get()
    if request.method == 'POST':
        cfg.prefix             = request.form.get('prefix', 'RTS').strip()[:10]
        cfg.separator          = request.form.get('separator', '-')[:3]
        cfg.use_category_code  = 'use_category_code' in request.form
        cfg.category_code_len  = int(request.form.get('category_code_len', 1))
        cfg.use_year           = 'use_year' in request.form
        cfg.year_format        = request.form.get('year_format', 'YY')
        cfg.consecutive_digits = int(request.form.get('consecutive_digits', 3))
        next_val               = request.form.get('next_consecutive', '').strip()
        if next_val.isdigit():
            cfg.next_consecutive = int(next_val)
        log_action('update', 'id_config', entity_name='ID Configuration')
        db.session.commit()
        flash('Configuración de ID guardada.', 'success')
        return redirect(url_for('id_config_view'))
    return render_template('setup/id_config.html', cfg=cfg)


@app.route('/setup/id-config/generate')
@admin_required
def id_config_generate():
    category = request.args.get('category', '')
    cfg = IDConfig.get()
    tag = cfg.generate_tag(category_name=category)
    cfg.next_consecutive += 1
    db.session.commit()
    return jsonify({'tag': tag})


# ── Labels ────────────────────────────────────────────────────────────────────

@app.route('/setup/labels')
@admin_required
def labels_index():
    assets = Asset.query.order_by(Asset.asset_tag).all()
    return render_template('setup/labels.html', assets=assets)


@app.route('/setup/labels/print', methods=['POST'])
@admin_required
def labels_print():
    ids = request.form.getlist('asset_ids')
    if not ids:
        flash('Selecciona al menos un activo para imprimir.', 'warning')
        return redirect(url_for('labels_index'))
    assets = Asset.query.filter(Asset.id.in_([int(i) for i in ids])).all()
    labels = []
    for a in assets:
        labels.append({
            'asset':  a,
            'qr_b64': _qr_b64(a.asset_tag),
        })
    return render_template('setup/labels_print.html', labels=labels)


@app.route('/setup/labels/single/<int:asset_id>')
@admin_required
def label_single(asset_id):
    a = Asset.query.get_or_404(asset_id)
    labels = [{'asset': a, 'qr_b64': _qr_b64(a.asset_tag)}]
    return render_template('setup/labels_print.html', labels=labels)


@app.errorhandler(403)
def forbidden(e):
    return render_template('errors/403.html'), 403

@app.errorhandler(404)
def not_found(e):
    return render_template('errors/404.html'), 404

@app.errorhandler(500)
def server_error(e):
    return render_template('errors/500.html'), 500

@app.errorhandler(429)
def rate_limited(e):
    return render_template('errors/429.html'), 429


# ── Auto-tracking scheduler desactivado ───────────────────────────────────────
# El plan gratuito de AfterShip permite 100 requests/día.
# Con pocos envíos es más eficiente rastrear manualmente con el botón
# "Rastrear ahora" en cada envío → 1 request por acción del usuario.
# Si en el futuro se tienen muchos envíos activos, reactivar el scheduler.

if __name__ == '__main__':
    app.run(debug=os.environ.get('FLASK_DEBUG', '0') == '1', port=5050, host='0.0.0.0')
