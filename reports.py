"""
RTS Excel Report Generator
Routes:
  /reports/assets.xlsx       → Asset inventory (supports same filters as /assets)
  /reports/assignments.xlsx  → Assignments list
  /reports/shipments.xlsx    → Shipments list
  /reports/employees.xlsx    → Employee directory
  /reports/dashboard.xlsx    → Multi-sheet dashboard summary
"""
import io
from datetime import datetime, date

from flask import Blueprint, request, send_file, session, redirect, url_for, flash
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

from models import db, Asset, Assignment, Shipment, Employee, Category, Client, Department
from auth import login_required, module_required

reports_bp = Blueprint('reports', __name__, url_prefix='/reports')

# ─────────────────────────────────────────────────────────────
# Brand constants
# ─────────────────────────────────────────────────────────────
NAVY    = '233C6E'
BLUE    = '089ACF'
LBLUE   = 'EBF5FB'   # light blue alternating row
LGRAY   = 'F4F6FA'   # header band bg
WHITE   = 'FFFFFF'
DKGRAY  = '54595F'
GREEN   = '28A745'
RED     = 'DC3545'
ORANGE  = 'FD7E14'
AMBER   = 'FFC107'

STATUS_COLORS = {
    'available':   '28A745',
    'in_use':      '0076D7',
    'maintenance': 'FFA000',
    'retired':     '6C757D',
    'disposed':    'DC3545',
    # assignments
    'active':      '28A745',
    'returned':    '6C757D',
    # shipments
    'pendiente':   'FFA000',
    'en_transito': '0076D7',
    'en_aduana':   '6F42C1',
    'entregado':   '28A745',
    'devuelto':    '6C757D',
}

PRIORITY_COLORS = {
    'critical': 'DC3545',
    'high':     'FD7E14',
    'medium':   'FFC107',
    'low':      '28A745',
}


# ─────────────────────────────────────────────────────────────
# Style helpers
# ─────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def _font(bold=False, color=WHITE, size=11, italic=False):
    return Font(name='Calibri', bold=bold, color=color, size=size, italic=italic)


def _border():
    side = Side(style='thin', color='D0D7E8')
    return Border(left=side, right=side, top=side, bottom=side)


def _center(wrap=False):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)


def _left(wrap=False):
    return Alignment(horizontal='left', vertical='center', wrap_text=wrap)


def _right():
    return Alignment(horizontal='right', vertical='center')


def _status_font(hex_color):
    """Bold colored font for status cells."""
    return Font(name='Calibri', bold=True, color=hex_color, size=10)


# ─────────────────────────────────────────────────────────────
# Workbook builder
# ─────────────────────────────────────────────────────────────

def _make_wb():
    return Workbook()


def _add_rts_header(ws, title: str, subtitle: str = '', total: int | None = None,
                    n_cols: int = 10):
    """
    Inserts a 5-row branded header into worksheet ws.
    Returns the row number where data headers should start (row 6).
    """
    # Row 1 — Company name
    ws.row_dimensions[1].height = 32
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(1, 1,
                value='Remote Team Solutions  ·  RTS Asset Management')
    c.fill      = _fill(NAVY)
    c.font      = Font(name='Calibri', bold=True, color=WHITE, size=16)
    c.alignment = _center()

    # Row 2 — Report title
    ws.row_dimensions[2].height = 22
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    c = ws.cell(2, 1, value=title)
    c.fill      = _fill(BLUE)
    c.font      = Font(name='Calibri', bold=True, color=WHITE, size=13)
    c.alignment = _center()

    # Row 3 — Subtitle / filter info
    if subtitle:
        ws.row_dimensions[3].height = 16
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols)
        c = ws.cell(3, 1, value=subtitle)
        c.fill      = _fill(LGRAY)
        c.font      = Font(name='Calibri', italic=True, color=DKGRAY, size=10)
        c.alignment = _center()

    # Row 4 — Meta line (generated date + count)
    ws.row_dimensions[4].height = 15
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=n_cols)
    meta_parts = [f'Generated: {datetime.now().strftime("%B %d, %Y  %H:%M")}']
    if total is not None:
        meta_parts.append(f'Total records: {total}')
    c = ws.cell(4, 1, value='   ·   '.join(meta_parts))
    c.fill      = _fill(LGRAY)
    c.font      = Font(name='Calibri', color='8A9BB5', size=9, italic=True)
    c.alignment = _center()

    # Row 5 — blank spacer
    ws.row_dimensions[5].height = 4
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=n_cols)
    ws.cell(5, 1).fill = _fill(LGRAY)

    return 6   # header row


def _write_col_headers(ws, headers, row, col_widths=None):
    """Write column header row with navy styling."""
    ws.row_dimensions[row].height = 20
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row, ci, value=h)
        c.fill      = _fill(NAVY)
        c.font      = Font(name='Calibri', bold=True, color=WHITE, size=10)
        c.alignment = _center(wrap=True)
        c.border    = _border()
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = ws.cell(row + 1, 1)


def _data_row(ws, row, values, alt=False, col_formats=None):
    """Write a data row with alternating background."""
    bg = LBLUE if alt else WHITE
    ws.row_dimensions[row].height = 16
    for ci, val in enumerate(values, 1):
        c  = ws.cell(row, ci, value=val)
        c.fill      = _fill(bg)
        c.font      = Font(name='Calibri', color='2C3E50', size=10)
        c.alignment = _left(wrap=False)
        c.border    = _border()
        # Apply overrides from col_formats: {col_idx: {'color': hex, 'bold': bool, 'align': 'center'}}
        if col_formats and ci in col_formats:
            fmt = col_formats[ci]
            if 'color' in fmt:
                c.font = Font(name='Calibri', bold=fmt.get('bold', False),
                              color=fmt['color'], size=10)
            if fmt.get('align') == 'center':
                c.alignment = _center()
            elif fmt.get('align') == 'right':
                c.alignment = _right()
    return row + 1


def _add_footer(ws, row, n_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row, 1,
                value=f'Remote Team Solutions  ·  RTS Intranet  ·  © {date.today().year}  ·  Confidential')
    c.fill      = _fill(NAVY)
    c.font      = Font(name='Calibri', color='8AB0D0', size=9, italic=True)
    c.alignment = _center()
    ws.row_dimensions[row].height = 14


def _send(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


def _today_str():
    return date.today().strftime('%Y%m%d')


# ─────────────────────────────────────────────────────────────
# Routes
# ─────────────────────────────────────────────────────────────

@reports_bp.route('/assets.xlsx')
@module_required('inventory')
def export_assets():
    # Respect same filters as /assets
    q               = request.args.get('q', '')
    status_filter   = request.args.get('status', '')
    category_filter = request.args.get('category', '')
    loc_filter      = request.args.get('location_type', '')

    # Department scope (same logic as inventory_dashboard)
    from models import User as _User
    sess_user = session.get('user', {})
    is_admin  = sess_user.get('role') == 'admin'
    db_user   = _User.query.get(sess_user.get('id')) if sess_user.get('id') else None
    dept_filter = None
    if not is_admin and db_user and db_user.department_id:
        dept_filter = db_user.department_id

    query = Asset.query
    if dept_filter:
        query = query.filter_by(department_id=dept_filter)
    if q:
        like = f'%{q}%'
        query = query.outerjoin(Client, Asset.client_id == Client.id).filter(db.or_(
            Asset.name.ilike(like), Asset.asset_tag.ilike(like),
            Asset.serial_number.ilike(like), Asset.manufacturer.ilike(like),
            Asset.model.ilike(like),
        ))
    if status_filter:
        query = query.filter_by(status=status_filter)
    if category_filter:
        query = query.filter_by(category_id=category_filter)
    if loc_filter:
        query = query.filter_by(location_type=loc_filter)
    assets = query.order_by(Asset.asset_tag).all()

    # Build subtitle from active filters
    parts = []
    if status_filter:   parts.append(f'Status: {status_filter.replace("_"," ").title()}')
    if category_filter:
        cat = Category.query.get(category_filter)
        if cat: parts.append(f'Category: {cat.name}')
    if loc_filter:      parts.append(f'Location: {loc_filter.replace("_"," ").title()}')
    if q:               parts.append(f'Search: "{q}"')
    subtitle = '  ·  '.join(parts) if parts else 'All Assets'

    COLS = [
        'Asset Tag', 'Name', 'Category', 'Status', 'Location Type', 'Location',
        'Manufacturer', 'Model', 'CPU', 'RAM', 'OS Version', 'Serial Number',
        'Client', 'Department', 'Assigned To', 'Employee ID',
        'Purchase Date', 'Purchase Cost (MXN)', 'Supplier', 'Warranty Expiry',
        'Last Maintenance', 'Notes',
    ]
    WIDTHS = [
        14, 28, 16, 14, 14, 20,
        16, 20, 26, 10, 18, 20,
        18, 16, 22, 14,
        14, 18, 20, 15,
        15, 32,
    ]
    N = len(COLS)

    wb = _make_wb()
    ws = wb.active
    ws.title = 'Assets'

    hdr_row = _add_rts_header(ws, 'Asset Inventory Report', subtitle, len(assets), N)
    _write_col_headers(ws, COLS, hdr_row, WIDTHS)

    STATUS_LABELS = {
        'available': 'Available', 'in_use': 'In Use',
        'maintenance': 'Maintenance', 'retired': 'Retired', 'disposed': 'Disposed',
    }
    LOC_LABELS = {
        'en_sitio': 'On-Site', 'hibrido': 'Hybrid', 'foraneo': 'Remote',
    }

    for i, a in enumerate(assets):
        curr = a.current_assignment
        emp  = curr.employee if curr else None
        dept = Department.query.get(a.department_id) if a.department_id else None
        status_color = STATUS_COLORS.get(a.status, DKGRAY)

        row = hdr_row + 1 + i
        vals = [
            a.asset_tag,
            a.name,
            a.category.name if a.category else '',
            STATUS_LABELS.get(a.status, a.status),
            LOC_LABELS.get(a.location_type, a.location_type or ''),
            a.location or '',
            a.manufacturer or '',
            a.model or '',
            a.cpu or '',
            a.ram or '',
            a.os_version or '',
            a.serial_number or '',
            a.client.name if a.client else '',
            dept.name if dept else '',
            emp.name if emp else '',
            emp.employee_id if emp else '',
            a.purchase_date.strftime('%Y-%m-%d') if a.purchase_date else '',
            a.purchase_cost if a.purchase_cost else '',
            a.supplier or '',
            a.warranty_expiry.strftime('%Y-%m-%d') if a.warranty_expiry else '',
            a.last_maintenance.strftime('%Y-%m-%d') if a.last_maintenance else '',
            a.notes or '',
        ]
        col_fmts = {
            4: {'color': status_color, 'bold': True, 'align': 'center'},
            18: {'align': 'right'},
        }
        _data_row(ws, row, vals, alt=(i % 2 == 1), col_formats=col_fmts)

    _add_footer(ws, hdr_row + 1 + len(assets) + 1, N)
    return _send(wb, f'RTS_Assets_{_today_str()}.xlsx')


@reports_bp.route('/assignments.xlsx')
@module_required('inventory')
def export_assignments():
    active_only = request.args.get('active', '1')
    query = Assignment.query
    if active_only == '1':
        query = query.filter_by(returned_date=None)
    items = query.order_by(Assignment.assigned_date.desc()).all()

    subtitle = 'Active Assignments Only' if active_only == '1' else 'All Assignments (including returned)'

    COLS = [
        'Asset Tag', 'Asset Name', 'Category', 'Status',
        'Employee Name', 'Employee ID', 'Department', 'Email',
        'Assigned Date', 'Returned Date', 'Days Assigned', 'Notes',
    ]
    WIDTHS = [14, 28, 16, 14, 24, 14, 20, 28, 14, 14, 14, 30]
    N = len(COLS)

    wb = _make_wb()
    ws = wb.active
    ws.title = 'Assignments'

    hdr_row = _add_rts_header(ws, 'Assignments Report', subtitle, len(items), N)
    _write_col_headers(ws, COLS, hdr_row, WIDTHS)

    today = date.today()
    for i, a in enumerate(items):
        is_active = a.returned_date is None
        days = (a.returned_date or today) - a.assigned_date if a.assigned_date else None
        row  = hdr_row + 1 + i
        vals = [
            a.asset.asset_tag if a.asset else '',
            a.asset.name if a.asset else '',
            a.asset.category.name if a.asset and a.asset.category else '',
            a.asset.status.replace('_', ' ').title() if a.asset else '',
            a.employee.name if a.employee else '',
            a.employee.employee_id if a.employee else '',
            a.employee.department if a.employee else '',
            a.employee.email if a.employee else '',
            a.assigned_date.strftime('%Y-%m-%d') if a.assigned_date else '',
            a.returned_date.strftime('%Y-%m-%d') if a.returned_date else '',
            days.days if days else '',
            a.notes or '',
        ]
        col_fmts = {
            4:  {'color': '0076D7' if is_active else DKGRAY, 'bold': True, 'align': 'center'},
            11: {'align': 'right'},
        }
        _data_row(ws, row, vals, alt=(i % 2 == 1), col_formats=col_fmts)

    _add_footer(ws, hdr_row + 1 + len(items) + 1, N)
    return _send(wb, f'RTS_Assignments_{_today_str()}.xlsx')


@reports_bp.route('/shipments.xlsx')
@module_required('inventory')
def export_shipments():
    items = Shipment.query.order_by(Shipment.created_at.desc()).all()

    COLS = [
        'Tracking #', 'Carrier', 'Asset Tag', 'Asset Name',
        'Origin', 'Destination', 'Recipient',
        'Status', 'Ship Date', 'Est. Delivery', 'Actual Delivery', 'Notes',
    ]
    WIDTHS = [20, 10, 14, 26, 24, 24, 20, 14, 14, 14, 14, 30]
    N = len(COLS)

    STATUS_LABELS = {
        'pendiente': 'Pending', 'en_transito': 'In Transit',
        'en_aduana': 'In Customs', 'entregado': 'Delivered', 'devuelto': 'Returned',
    }

    wb = _make_wb()
    ws = wb.active
    ws.title = 'Shipments'

    hdr_row = _add_rts_header(ws, 'Shipments Report', 'All Shipments', len(items), N)
    _write_col_headers(ws, COLS, hdr_row, WIDTHS)

    for i, s in enumerate(items):
        row = hdr_row + 1 + i
        status_color = STATUS_COLORS.get(s.status, DKGRAY)
        vals = [
            s.tracking_number,
            s.carrier,
            s.asset.asset_tag if s.asset else '',
            s.asset.name if s.asset else '',
            s.origin or '',
            s.destination or '',
            s.recipient_name or '',
            STATUS_LABELS.get(s.status, s.status),
            s.ship_date.strftime('%Y-%m-%d') if s.ship_date else '',
            s.estimated_delivery.strftime('%Y-%m-%d') if s.estimated_delivery else '',
            s.actual_delivery.strftime('%Y-%m-%d') if s.actual_delivery else '',
            s.notes or '',
        ]
        col_fmts = {8: {'color': status_color, 'bold': True, 'align': 'center'}}
        _data_row(ws, row, vals, alt=(i % 2 == 1), col_formats=col_fmts)

    _add_footer(ws, hdr_row + 1 + len(items) + 1, N)
    return _send(wb, f'RTS_Shipments_{_today_str()}.xlsx')


@reports_bp.route('/employees.xlsx')
@module_required('inventory')
def export_employees():
    items = Employee.query.order_by(Employee.name).all()

    COLS = [
        'Employee ID', 'Name', 'Department', 'Email', 'Phone',
        'Status', 'Total Assets', 'Current Asset Tags',
    ]
    WIDTHS = [14, 26, 20, 30, 16, 10, 13, 40]
    N = len(COLS)

    wb = _make_wb()
    ws = wb.active
    ws.title = 'Employees'

    hdr_row = _add_rts_header(ws, 'Employee Directory', 'All Employees', len(items), N)
    _write_col_headers(ws, COLS, hdr_row, WIDTHS)

    for i, e in enumerate(items):
        active_assigns = [a for a in e.assignments if not a.returned_date]
        tags = ', '.join(a.asset.asset_tag for a in active_assigns if a.asset)
        row  = hdr_row + 1 + i
        vals = [
            e.employee_id,
            e.name,
            e.department or '',
            e.email or '',
            e.phone or '',
            'Active' if e.active else 'Inactive',
            len(active_assigns),
            tags,
        ]
        col_fmts = {
            6:  {'color': GREEN if e.active else DKGRAY, 'bold': True, 'align': 'center'},
            7:  {'align': 'center'},
        }
        _data_row(ws, row, vals, alt=(i % 2 == 1), col_formats=col_fmts)

    _add_footer(ws, hdr_row + 1 + len(items) + 1, N)
    return _send(wb, f'RTS_Employees_{_today_str()}.xlsx')


@reports_bp.route('/dashboard.xlsx')
@module_required('inventory')
def export_dashboard():
    """Multi-sheet executive summary report."""
    # Department scope
    from models import User as _User
    sess_user = session.get('user', {})
    is_admin  = sess_user.get('role') == 'admin'
    db_user   = _User.query.get(sess_user.get('id')) if sess_user.get('id') else None
    dept_filter = None
    if not is_admin and db_user and db_user.department_id:
        dept_filter = db_user.department_id

    def _asset_q():
        q = Asset.query
        if dept_filter:
            q = q.filter_by(department_id=dept_filter)
        return q

    today = date.today()
    wb    = _make_wb()

    # ── Sheet 1: Executive Summary ───────────────────────────
    ws1 = wb.active
    ws1.title = '📊 Summary'
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 26

    # Header
    hdr = _add_rts_header(ws1, 'Inventory Dashboard — Executive Summary',
                          f'Report date: {today.strftime("%B %d, %Y")}',
                          None, 3)

    # KPI title row
    r = hdr
    ws1.row_dimensions[r].height = 18
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws1.cell(r, 1, value='KEY PERFORMANCE INDICATORS')
    c.fill      = _fill(LGRAY)
    c.font      = Font(name='Calibri', bold=True, color=NAVY, size=11)
    c.alignment = _left()
    r += 1

    def _kpi(label, val, color=NAVY, pct=None):
        nonlocal r
        ws1.row_dimensions[r].height = 20
        c1 = ws1.cell(r, 1, value=label)
        c1.fill  = _fill(WHITE if r % 2 == 0 else LBLUE)
        c1.font  = Font(name='Calibri', color=DKGRAY, size=11)
        c1.alignment = _left()
        c1.border    = _border()

        c2 = ws1.cell(r, 2, value=val)
        c2.fill  = _fill(WHITE if r % 2 == 0 else LBLUE)
        c2.font  = Font(name='Calibri', bold=True, color=color, size=14)
        c2.alignment = _center()
        c2.border    = _border()

        c3 = ws1.cell(r, 3, value=f'{pct:.1f}%' if pct is not None else '')
        c3.fill  = _fill(WHITE if r % 2 == 0 else LBLUE)
        c3.font  = Font(name='Calibri', color=DKGRAY, size=10, italic=True)
        c3.alignment = _center()
        c3.border    = _border()
        r += 1

    total      = _asset_q().count()
    available  = _asset_q().filter_by(status='available').count()
    in_use     = _asset_q().filter_by(status='in_use').count()
    maint      = _asset_q().filter_by(status='maintenance').count()
    retired    = _asset_q().filter_by(status='retired').count()
    disposed   = _asset_q().filter_by(status='disposed').count()
    foraneo    = _asset_q().filter_by(location_type='foraneo').count()
    hibrido    = _asset_q().filter_by(location_type='hibrido').count()
    active_emp = Employee.query.filter_by(active=True).count()
    in_transit = Shipment.query.filter_by(status='en_transito').count()
    active_asg = Assignment.query.filter_by(returned_date=None).count()

    pct = lambda n: (n / total * 100) if total else 0

    ws1.cell(r, 1, value='Metric').font = Font(name='Calibri', bold=True, color=WHITE, size=10)
    ws1.cell(r, 1).fill = _fill(NAVY)
    ws1.cell(r, 2, value='Value').font  = Font(name='Calibri', bold=True, color=WHITE, size=10)
    ws1.cell(r, 2).fill = _fill(NAVY)
    ws1.cell(r, 2).alignment = _center()
    ws1.cell(r, 3, value='% of Total Assets').font = Font(name='Calibri', bold=True, color=WHITE, size=10)
    ws1.cell(r, 3).fill = _fill(NAVY)
    ws1.cell(r, 3).alignment = _center()
    ws1.row_dimensions[r].height = 18
    r += 1

    _kpi('Total Assets',              total,      NAVY)
    _kpi('Available',                 available,  GREEN,  pct(available))
    _kpi('In Use',                    in_use,     BLUE,   pct(in_use))
    _kpi('Under Maintenance',         maint,      ORANGE, pct(maint))
    _kpi('Retired',                   retired,    DKGRAY, pct(retired))
    _kpi('Disposed',                  disposed,   RED,    pct(disposed))
    _kpi('Remote / Foráneo',          foraneo,    '6F42C1', pct(foraneo))
    _kpi('Hybrid',                    hibrido,    BLUE,   pct(hibrido))
    _kpi('Active Employees',          active_emp, NAVY)
    _kpi('Active Assignments',        active_asg, BLUE)
    _kpi('Shipments in Transit',      in_transit, AMBER)

    r += 1
    _add_footer(ws1, r, 3)

    # ── Sheet 2: By Category ────────────────────────────────
    ws2 = wb.create_sheet('📦 By Category')
    ws2.column_dimensions['A'].width = 26
    for col in ['B', 'C', 'D', 'E', 'F']: ws2.column_dimensions[col].width = 14

    h2 = _add_rts_header(ws2, 'Assets by Category', None, None, 6)
    _write_col_headers(ws2,
        ['Category', 'Total', 'Available', 'In Use', 'Maintenance', 'Retired/Disposed'],
        h2, None)

    cats = Category.query.order_by(Category.name).all()
    for i, cat in enumerate(cats):
        base = _asset_q().filter_by(category_id=cat.id)
        vals = [
            cat.name,
            base.count(),
            base.filter_by(status='available').count(),
            base.filter_by(status='in_use').count(),
            base.filter_by(status='maintenance').count(),
            base.filter(Asset.status.in_(['retired', 'disposed'])).count(),
        ]
        col_fmts = {
            2: {'align': 'center', 'color': NAVY, 'bold': True},
            3: {'align': 'center', 'color': GREEN},
            4: {'align': 'center', 'color': BLUE},
            5: {'align': 'center', 'color': ORANGE},
            6: {'align': 'center', 'color': DKGRAY},
        }
        _data_row(ws2, h2 + 1 + i, vals, alt=(i % 2 == 1), col_formats=col_fmts)

    _add_footer(ws2, h2 + 1 + len(cats) + 1, 6)

    # ── Sheet 3: By Department ──────────────────────────────
    ws3 = wb.create_sheet('🏢 By Department')
    ws3.column_dimensions['A'].width = 22
    ws3.column_dimensions['B'].width = 10
    for col in ['C', 'D', 'E', 'F', 'G']: ws3.column_dimensions[col].width = 14

    h3 = _add_rts_header(ws3, 'Assets by Department', None, None, 7)
    _write_col_headers(ws3,
        ['Department', 'Code', 'Total Assets', 'Available', 'In Use', 'Maintenance', 'Users'],
        h3, None)

    depts = Department.query.filter_by(active=True).order_by(Department.name).all()
    for i, d in enumerate(depts):
        base = Asset.query.filter_by(department_id=d.id)
        usr  = len([u for u in d.users if u.active])
        vals = [
            d.name, d.code,
            base.count(),
            base.filter_by(status='available').count(),
            base.filter_by(status='in_use').count(),
            base.filter_by(status='maintenance').count(),
            usr,
        ]
        col_fmts = {
            2: {'align': 'center'},
            3: {'align': 'center', 'color': NAVY, 'bold': True},
            4: {'align': 'center', 'color': GREEN},
            5: {'align': 'center', 'color': BLUE},
            6: {'align': 'center', 'color': ORANGE},
            7: {'align': 'center'},
        }
        _data_row(ws3, h3 + 1 + i, vals, alt=(i % 2 == 1), col_formats=col_fmts)

    # Unassigned row
    unassigned = _asset_q().filter(Asset.department_id == None).count()  # noqa: E711
    if unassigned:
        idx = len(depts)
        vals = ['— Unassigned —', '', unassigned, '', '', '', '']
        col_fmts = {3: {'align': 'center', 'color': DKGRAY, 'bold': True}}
        _data_row(ws3, h3 + 1 + idx, vals, alt=(idx % 2 == 1), col_formats=col_fmts)

    _add_footer(ws3, h3 + 2 + len(depts) + 1, 7)

    # ── Sheet 4: Recent Assignments ──────────────────────────
    ws4 = wb.create_sheet('📋 Assignments')
    ws4.column_dimensions['A'].width = 14
    ws4.column_dimensions['B'].width = 28
    ws4.column_dimensions['C'].width = 16
    ws4.column_dimensions['D'].width = 22
    ws4.column_dimensions['E'].width = 14
    ws4.column_dimensions['F'].width = 14
    ws4.column_dimensions['G'].width = 14
    ws4.column_dimensions['H'].width = 14

    recent_asgns = (Assignment.query
                    .filter_by(returned_date=None)
                    .order_by(Assignment.assigned_date.desc())
                    .limit(100).all())

    h4 = _add_rts_header(ws4, 'Active Assignments (last 100)',
                         f'As of {today.strftime("%B %d, %Y")}',
                         len(recent_asgns), 8)
    _write_col_headers(ws4,
        ['Asset Tag', 'Asset Name', 'Category', 'Employee', 'Dept.', 'Assigned', 'Days', 'Location'],
        h4, None)

    for i, a in enumerate(recent_asgns):
        days = (today - a.assigned_date).days if a.assigned_date else ''
        vals = [
            a.asset.asset_tag if a.asset else '',
            a.asset.name if a.asset else '',
            a.asset.category.name if a.asset and a.asset.category else '',
            a.employee.name if a.employee else '',
            a.employee.department if a.employee else '',
            a.assigned_date.strftime('%Y-%m-%d') if a.assigned_date else '',
            days,
            a.asset.location_type.replace('_', ' ').title() if a.asset else '',
        ]
        col_fmts = {7: {'align': 'right'}}
        _data_row(ws4, h4 + 1 + i, vals, alt=(i % 2 == 1), col_formats=col_fmts)

    _add_footer(ws4, h4 + 1 + len(recent_asgns) + 1, 8)

    # ── Sheet 5: Shipments in Transit ───────────────────────
    ws5 = wb.create_sheet('🚚 Shipments')
    for ci, w in enumerate([20, 10, 14, 26, 22, 22, 18, 14, 14, 14], 1):
        ws5.column_dimensions[get_column_letter(ci)].width = w

    ships = (Shipment.query
             .filter(Shipment.status.notin_(['entregado', 'devuelto']))
             .order_by(Shipment.created_at.desc()).all())

    STATUS_LBL = {
        'pendiente': 'Pending', 'en_transito': 'In Transit',
        'en_aduana': 'In Customs', 'entregado': 'Delivered', 'devuelto': 'Returned',
    }

    h5 = _add_rts_header(ws5, 'Active Shipments', 'Excluding Delivered & Returned',
                         len(ships), 10)
    _write_col_headers(ws5,
        ['Tracking #', 'Carrier', 'Asset Tag', 'Asset Name',
         'Origin', 'Destination', 'Recipient', 'Status', 'Ship Date', 'Est. Delivery'],
        h5, None)

    for i, s in enumerate(ships):
        sc = STATUS_COLORS.get(s.status, DKGRAY)
        vals = [
            s.tracking_number, s.carrier,
            s.asset.asset_tag if s.asset else '',
            s.asset.name if s.asset else '',
            s.origin or '', s.destination or '', s.recipient_name or '',
            STATUS_LBL.get(s.status, s.status),
            s.ship_date.strftime('%Y-%m-%d') if s.ship_date else '',
            s.estimated_delivery.strftime('%Y-%m-%d') if s.estimated_delivery else '',
        ]
        col_fmts = {8: {'color': sc, 'bold': True, 'align': 'center'}}
        _data_row(ws5, h5 + 1 + i, vals, alt=(i % 2 == 1), col_formats=col_fmts)

    _add_footer(ws5, h5 + 1 + len(ships) + 1, 10)

    return _send(wb, f'RTS_Dashboard_{_today_str()}.xlsx')
